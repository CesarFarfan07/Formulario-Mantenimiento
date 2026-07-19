"""RACS router — form, CRUD, workers, dashboard, export."""
import os, uuid, io, datetime as dt_mod
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Request, Body
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func as _sf
import openpyxl
from openpyxl.styles import Font as XlFont, Alignment as XlAlign, PatternFill as XlFill, Border as XlBorder, Side as XlSide
from openpyxl.utils import get_column_letter, column_index_from_string

from ..core.database import get_db
from ..core.config import settings, BASE_DIR
from ..models import RacsReport, Guardia, RacsWorker, WorkerGuardia
from ..schemas import RacsReportCreate, RacsReportResponse, RacsWorkerCreate, RacsWorkerUpdate
from ..services.racs_service import (
    get_racs_period, is_guardia_on_site, get_worker_guardias,
    get_worker_cargos, get_all_workers,
)

PROJECT_DIR = os.path.dirname(BASE_DIR)
STATIC_DIR = os.path.join(PROJECT_DIR, "static")
TEMPLATE_DIR = os.path.join(BASE_DIR, "app", "templates")
templates = Jinja2Templates(directory=TEMPLATE_DIR)

router = APIRouter(tags=["RACS"])


# ─── Page routes ────────────────────────────────────────────────────────


@router.get("/racs", response_class=HTMLResponse)
def racs_form(request: Request):
    return templates.TemplateResponse("racs_form.html", {"request": request})


@router.get("/racs/dashboard", response_class=HTMLResponse)
def racs_dashboard_page(request: Request):
    return templates.TemplateResponse("racs_dashboard.html", {"request": request})


# ─── API: Period ────────────────────────────────────────────────────────


@router.get("/api/racs/period")
def racs_get_period():
    s, e = get_racs_period()
    return {"period_start": s.isoformat(), "period_end": e.isoformat()}


# ─── API: Workers ───────────────────────────────────────────────────────


@router.get("/api/racs/workers")
def racs_get_workers(db: Session = Depends(get_db)):
    wg_map = get_worker_guardias(db)
    cargos = get_worker_cargos(db)
    all_workers = get_all_workers(db)
    teams = {}
    for w in all_workers:
        teams.setdefault(w["group_name"], []).append(w)
    result = []
    for group, members in teams.items():
        workers_list = []
        for w in members:
            guardia_name = wg_map.get((w["name"], group), "")
            on_site = is_guardia_on_site(guardia_name, db=db)
            workers_list.append({"name": w["name"], "guardia": guardia_name, "on_site": on_site, "cargo": w["cargo"] or ""})
        result.append({"group": group, "workers": workers_list})
    return result


@router.get("/api/racs/workers/list")
def racs_workers_list(db: Session = Depends(get_db)):
    wg = get_worker_guardias(db)
    all_workers = get_all_workers(db)
    teams = {}
    for w in all_workers:
        teams.setdefault(w["group_name"], []).append(w)
    result = []
    for group, members in teams.items():
        workers_list = []
        for w in members:
            guardia_name = wg.get((w["name"], group), "")
            workers_list.append({"name": w["name"], "cargo": w["cargo"] or "", "guardia": guardia_name, "group": group})
        result.append({"group": group, "workers": workers_list})
    return result


@router.post("/api/racs/workers/create")
def racs_worker_create(data: RacsWorkerCreate, db: Session = Depends(get_db)):
    existing = db.query(RacsWorker).filter(RacsWorker.name == data.name, RacsWorker.group_name == data.group_name).first()
    if existing:
        if not existing.active:
            existing.active = True
            existing.cargo = data.cargo or existing.cargo
            db.commit()
            return {"ok": True, "id": existing.id}
        raise HTTPException(400, "El trabajador ya existe en este equipo")
    obj = RacsWorker(name=data.name, group_name=data.group_name, cargo=data.cargo, active=True)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return {"ok": True, "id": obj.id}


@router.put("/api/racs/workers/{worker_id}")
def racs_worker_update(worker_id: int, data: RacsWorkerUpdate, db: Session = Depends(get_db)):
    obj = db.query(RacsWorker).filter(RacsWorker.id == worker_id).first()
    if not obj:
        raise HTTPException(404, "Trabajador no encontrado")
    old_name = obj.name
    old_group = obj.group_name
    if data.name is not None: obj.name = data.name
    if data.group_name is not None: obj.group_name = data.group_name
    if data.cargo is not None: obj.cargo = data.cargo
    if data.active is not None: obj.active = data.active
    # Sync WorkerGuardia if name or group changed
    if (data.name is not None or data.group_name is not None) and (old_name != obj.name or old_group != obj.group_name):
        wg = db.query(WorkerGuardia).filter(
            WorkerGuardia.worker_name == old_name,
            WorkerGuardia.group_name == old_group
        ).first()
        if wg:
            wg.worker_name = obj.name
            wg.group_name = obj.group_name
    if data.guardia is not None:
        existing_wg = db.query(WorkerGuardia).filter(
            WorkerGuardia.worker_name == obj.name,
            WorkerGuardia.group_name == obj.group_name,
        ).first()
        if data.guardia:
            guardia = db.query(Guardia).filter(Guardia.name == data.guardia).first()
            if not guardia:
                raise HTTPException(400, "Guardia no encontrada")
            if existing_wg:
                existing_wg.guardia_id = guardia.id
            else:
                db.add(WorkerGuardia(worker_name=obj.name, group_name=obj.group_name, guardia_id=guardia.id))
        elif existing_wg:
            db.delete(existing_wg)
    db.commit()
    db.refresh(obj)
    return {"ok": True, "id": obj.id, "name": obj.name, "group_name": obj.group_name, "cargo": obj.cargo}


@router.post("/api/racs/workers/change-group")
def racs_worker_change_group(data: dict = Body(...), db: Session = Depends(get_db)):
    name = data.get("name")
    old_group = data.get("old_group")
    new_group = data.get("new_group")
    if not name or not old_group or not new_group:
        raise HTTPException(400, "Faltan datos")
    obj = db.query(RacsWorker).filter(RacsWorker.name == name, RacsWorker.group_name == old_group, RacsWorker.active == True).first()
    if not obj:
        raise HTTPException(404, "Trabajador no encontrado")
    obj.group_name = new_group
    wg = db.query(WorkerGuardia).filter(WorkerGuardia.worker_name == name, WorkerGuardia.group_name == old_group).first()
    if wg:
        wg.group_name = new_group
    db.commit()
    return {"ok": True}


@router.get("/api/racs/groups")
def racs_get_groups(db: Session = Depends(get_db)):
    q = db.query(RacsWorker.group_name).filter(RacsWorker.active == True).distinct().order_by(RacsWorker.group_name).all()
    return [g[0] for g in q]


@router.delete("/api/racs/workers/delete-by-name")
def racs_worker_delete_by_name(name: str = Query(...), group: str = Query(...), db: Session = Depends(get_db)):
    obj = db.query(RacsWorker).filter(RacsWorker.name == name, RacsWorker.group_name == group, RacsWorker.active == True).first()
    if not obj:
        raise HTTPException(404, "Trabajador no encontrado")
    obj.active = False
    db.commit()
    return {"ok": True}


@router.delete("/api/racs/workers/{worker_id}")
def racs_worker_delete(worker_id: int, db: Session = Depends(get_db)):
    obj = db.query(RacsWorker).filter(RacsWorker.id == worker_id).first()
    if not obj:
        raise HTTPException(404)
    obj.active = False
    db.commit()
    return {"ok": True}


# ─── API: Reports CRUD ──────────────────────────────────────────────────


@router.post("/api/racs", response_model=RacsReportResponse)
def racs_create(data: RacsReportCreate, db: Session = Depends(get_db)):
    ps, pe = get_racs_period()
    fecha = None
    if data.fecha_reporte:
        try:
            fecha = dt_mod.datetime.strptime(data.fecha_reporte, "%Y-%m-%d").date()
        except:
            fecha = dt_mod.datetime.utcnow().date()
    r = RacsReport(
        worker_name=data.worker_name, group_name=data.group_name,
        categoria=data.categoria, tipo=data.tipo, turno=data.turno,
        descripcion=data.descripcion, ubicacion=data.ubicacion, referencia=data.referencia,
        nivel=data.nivel, fecha_reporte=fecha, riesgo=data.riesgo,
        accion_correctiva=data.accion_correctiva, tipo_descripcion=data.tipo_descripcion,
        period_start=ps, period_end=pe,
        created_at=dt_mod.datetime.now(dt_mod.timezone.utc),
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return r


@router.post("/api/racs/upload-photo/{racs_id}")
def racs_upload_photo(racs_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    r = db.query(RacsReport).filter(RacsReport.id == racs_id).first()
    if not r:
        raise HTTPException(404, "RACS no encontrado")
    ext = os.path.splitext(file.filename or "foto.jpg")[1] or ".jpg"
    fname = f"racs_{racs_id}_{uuid.uuid4().hex[:8]}{ext}"
    rel_dir = "racs_photos"
    dest_dir = os.path.join(STATIC_DIR, rel_dir)
    os.makedirs(dest_dir, exist_ok=True)
    dest = os.path.join(dest_dir, fname)
    with open(dest, "wb") as f:
        f.write(file.file.read())
    r.foto = f"/static/{rel_dir}/{fname}"
    db.commit()
    return {"ok": True, "path": r.foto}


@router.get("/api/racs/list")
def racs_list(db: Session = Depends(get_db)):
    ps, pe = get_racs_period()
    rows = db.query(RacsReport).filter(RacsReport.period_start == ps, RacsReport.period_end == pe).order_by(RacsReport.created_at.desc()).all()
    return [
        {
            "id": r.id, "worker_name": r.worker_name, "group_name": r.group_name,
            "tipo": r.tipo, "categoria": r.categoria, "turno": r.turno,
            "descripcion": r.descripcion, "ubicacion": r.ubicacion, "referencia": r.referencia,
            "nivel": r.nivel, "fecha_reporte": (r.fecha_reporte.isoformat() if r.fecha_reporte else None),
            "riesgo": r.riesgo, "accion_correctiva": r.accion_correctiva, "tipo_descripcion": r.tipo_descripcion,
            "foto": r.foto,
            "created_at": (r.created_at.isoformat() if r.created_at.tzinfo else r.created_at.isoformat() + "Z"),
        }
        for r in rows
    ]


@router.get("/api/racs/worker-status")
def racs_worker_status(worker_name: str, db: Session = Depends(get_db)):
    ps, pe = get_racs_period()
    cnt = db.query(_sf.count(RacsReport.id)).filter(
        RacsReport.worker_name == worker_name,
        RacsReport.period_start == ps, RacsReport.period_end == pe,
    ).scalar() or 0
    return {"count": cnt}


# ─── API: Dashboard ─────────────────────────────────────────────────────


@router.get("/api/racs/dashboard-data")
def racs_dashboard_data(db: Session = Depends(get_db)):
    ps, pe = get_racs_period()
    wg_map = get_worker_guardias(db)
    counts = db.query(
        RacsReport.worker_name, RacsReport.group_name,
        _sf.count(RacsReport.id).label("cnt"),
    ).filter(RacsReport.period_start == ps, RacsReport.period_end == pe).group_by(RacsReport.worker_name, RacsReport.group_name).all()
    count_map = {(r.worker_name, r.group_name): r.cnt for r in counts}
    all_workers = get_all_workers(db)
    teams_dict = {}
    for w in all_workers:
        teams_dict.setdefault(w["group_name"], []).append(w["name"])
    teams = []
    total_workers = 0; on_site_count = 0; complete = 0; partial = 0; missing = 0
    for group, workers in teams_dict.items():
        members = []
        for w_name in workers:
            total_workers += 1
            guardia_name = wg_map.get((w_name, group), "")
            on_site = is_guardia_on_site(guardia_name, db=db)
            c = count_map.get((w_name, group), 0)
            members.append({"name": w_name, "count": c, "guardia": guardia_name, "on_site": on_site})
            if on_site:
                on_site_count += 1
                if c >= 2: complete += 1
                elif c == 1: partial += 1
                else: missing += 1
        teams.append({"team": group, "workers": members})
    return {
        "period_start": ps.isoformat(), "period_end": pe.isoformat(),
        "teams": teams,
        "summary": {"total": total_workers, "on_site": on_site_count,
                     "on_rest": total_workers - on_site_count,
                     "complete": complete, "partial": partial, "missing": missing},
    }


@router.get("/api/racs/dashboard-kpi")
def racs_dashboard_kpi(db: Session = Depends(get_db)):
    ps, pe = get_racs_period()
    wg_map = get_worker_guardias(db)
    all_reports = db.query(RacsReport).filter(RacsReport.period_start == ps, RacsReport.period_end == pe).all()
    total = len(all_reports)
    cat_count = {"Seguridad y Salud Ocupacional": 0, "Medio Ambiente": 0}
    tipo_count = {"Acto Subestándar": 0, "Condición Subestándar": 0}
    riesgo_count = {"Alto": 0, "Medio": 0, "Bajo": 0}
    turno_count = {"DÍA": 0, "NOCHE": 0}
    nivel_count = {}; guardia_count = {}; daily_count = {}; worker_racs = {}
    for r in all_reports:
        cat = r.categoria or "Sin categoría"
        if cat in cat_count: cat_count[cat] += 1
        tp = r.tipo or "Sin tipo"
        if tp in tipo_count: tipo_count[tp] += 1
        rg = r.riesgo or "Sin riesgo"
        if rg in riesgo_count: riesgo_count[rg] += 1
        tn = r.turno or "Sin turno"
        if tn in turno_count: turno_count[tn] += 1
        nv = r.nivel or "Sin nivel"
        nivel_count[nv] = nivel_count.get(nv, 0) + 1
        g = wg_map.get((r.worker_name, r.group_name), "Sin guardia")
        guardia_count[g] = guardia_count.get(g, 0) + 1
        day_key = r.created_at.strftime("%Y-%m-%d")
        daily_count[day_key] = daily_count.get(day_key, 0) + 1
        wk = (r.worker_name, r.group_name)
        if wk not in worker_racs:
            worker_racs[wk] = {"count": 0, "cat": {}, "tipo": {}, "riesgo": {}, "guardia": g}
        worker_racs[wk]["count"] += 1
        worker_racs[wk]["cat"][cat] = worker_racs[wk]["cat"].get(cat, 0) + 1
        worker_racs[wk]["tipo"][tp] = worker_racs[wk]["tipo"].get(tp, 0) + 1
        worker_racs[wk]["riesgo"][rg] = worker_racs[wk]["riesgo"].get(rg, 0) + 1
    all_workers = get_all_workers(db)
    active_workers = 0
    for w in all_workers:
        g = wg_map.get((w["name"], w["group_name"]), "")
        if is_guardia_on_site(g, db=db):
            active_workers += 1
    daily_trend = [{"date": k, "count": v} for k, v in sorted(daily_count.items())]
    worker_detail = []
    for (name, group), data in worker_racs.items():
        g = data["guardia"]
        on_site = is_guardia_on_site(g, db=db)
        worker_detail.append({"name": name, "group": group, "count": data["count"], "guardia": g, "on_site": on_site, "riesgos": data["riesgo"]})
    nivel_sorted = [{"nivel": k, "count": v} for k, v in sorted(nivel_count.items(), key=lambda x: -x[1])]
    return {
        "total_racs": total, "active_workers": active_workers,
        "period_start": ps.isoformat(), "period_end": pe.isoformat(),
        "by_categoria": [{"categoria": k, "count": v} for k, v in cat_count.items()],
        "by_tipo": [{"tipo": k, "count": v} for k, v in tipo_count.items()],
        "by_riesgo": [{"riesgo": k, "count": v} for k, v in riesgo_count.items()],
        "by_turno": [{"turno": k, "count": v} for k, v in turno_count.items()],
        "by_nivel": nivel_sorted,
        "by_guardia": [{"guardia": k, "count": v} for k, v in sorted(guardia_count.items())],
        "daily_trend": daily_trend, "worker_detail": worker_detail,
    }


# ─── API: Excel export ──────────────────────────────────────────────────


@router.get("/api/racs/{racs_id}/excel")
def racs_download_excel(racs_id: int, db: Session = Depends(get_db)):
    from ..excel_export import build_excel
    r = db.query(RacsReport).filter(RacsReport.id == racs_id).first()
    if not r:
        raise HTTPException(404, "RACS no encontrado")
    TEMPLATE_PATH = os.path.join(TEMPLATE_DIR, "racs_template.xlsx")
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    riesgo_map = {"Alto": "ALTO", "Medio": "MEDIO", "Bajo": "BAJO"}
    checklist_map = {
        1: "1.- Métodos de Trabajo", 2: "2.- Orden y Limpieza",
        3: "3.- EPP (Casco, Lentes, Respirador, Arnés, etc.)",
        4: "4.- Bloqueo y señalización (LOTO)", 5: "5.- Herramientas Manuales",
        6: "6.- Herramientas Eléctricas", 7: "7.- Equipos Móviles (camioneta, volquete, etc.)",
        8: "8.- Izaje y Maniobras", 9: "9.- Trabajos en Altura",
        10: "10.- Espacios Confinados", 11: "11.- Instalaciones Eléctricas",
        12: "12.- Sustancias Peligrosas", 13: "13.- Protección Contra Incendios",
        14: "14.- Señalización de Seguridad", 15: "15.- Permisos de Trabajo",
        16: "16.- Guarda de Seguridad", 17: "17.- Sistemas de Ventilación",
        18: "18.- Sistemas de Bombeo", 19: "19.- Estabilidad de Terreno",
        20: "20.- Iluminación", 21: "21.- Vías de Acceso y Escape",
        22: "22.- Comunicaciones", 23: "23.- Primeros Auxilios",
        24: "24.- Equipos de Respiración", 25: "25.- Monitoreo de Gases",
        26: "26.- Control de Polución", 27: "27.- Manejo de Residuos",
        28: "28.- Uso de Agua", 29: "29.- Flora y Fauna",
        30: "30.- Ruido y Vibraciones", 31: "31.- Condición del Piso",
        32: "32.- Escaleras y Pasarelas", 33: "33.- Barandas y Rodapiés",
        34: "34.- Plataformas de Trabajo", 35: "35.- Sistema de Control",
        36: "36.- Equipo de Emergencia", 37: "37.- Ducha de Emergencia / Lavaojos",
        38: "38.- Almacenamiento", 39: "39.- Vigencia de Permisos",
        40: "40.- Uso de Radio Frecuencia", 41: "41.- Gestión de Tránsito",
        42: "42.- Otros",
    }
    ws["C8"] = r.worker_name or ""
    ws["C9"] = r.ubicacion or ""
    ws["C10"] = r.group_name or ""
    ws["G9"] = r.turno or ""
    ws["G10"] = riesgo_map.get(r.riesgo, "")
    desc_cell = ws["C14"]
    desc_cell.value = r.descripcion or ""
    acc_cell = ws["C15"]
    acc_cell.value = r.accion_correctiva or ""
    if r.tipo_descripcion and r.tipo_descripcion.isdigit():
        idx = int(r.tipo_descripcion)
        label = checklist_map.get(idx, "")
        cat_name = "SEGURIDAD" if r.categoria == "Seguridad y Salud Ocupacional" else "MEDIO AMBIENTE"
        for row in ws.iter_rows(min_row=17, max_row=58, min_col=1, max_col=3):
            cell = row[0]
            if cell.value and isinstance(cell.value, str):
                if label in cell.value:
                    row[2].value = "X"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"RACS_{r.worker_name.replace(' ','_')}_{r.created_at.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/api/racs/database-excel")
def racs_database_excel(db: Session = Depends(get_db)):
    reports = db.query(RacsReport).order_by(RacsReport.created_at.asc()).all()
    wg_map = get_worker_guardias(db)
    cargo_map = get_worker_cargos(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RACS GENERAL PROSOL SA"
    hdr_fill = XlFill("solid", fgColor="000000")
    hdr_font = XlFont(bold=True, color="FFFFFF", size=10, name="Calibri")
    hdr_align = XlAlign(horizontal="center", vertical="center", wrap_text=True)
    title_font = XlFont(bold=True, size=14, name="Calibri")
    sub_font = XlFont(size=8, name="Calibri")
    cell_font = XlFont(size=9, name="Calibri")
    cell_align = XlAlign(vertical="top", wrap_text=True)
    thin_side = XlSide(style="thin", color="000000")
    thin_border = XlBorder(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    ws.merge_cells("A1:B2")
    ws["A1"].font = title_font
    ws.merge_cells("C1:R2")
    ws["C1"].value = "BASE DE DATOS DE REPORTE DE ACTO Y CONDICIONES SUBESTANDARES - RACS CIA PROSOL ISPACAS S.A."
    ws["C1"].font = title_font
    ws["C1"].alignment = XlAlign(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("S1:S2")
    ws["S1"].font = sub_font
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    headers = [
        ("A", "CODIGO", 8), ("B", "SEMANA", 10), ("C", "MES", 12),
        ("D", "FECHA DE REPORTE", 14), ("E", "AREA", 10), ("F", "TIPO", 12),
        ("G", "GUARDIA", 10), ("H", "TURNO", 10), ("I", "NOMBRE DEL REPORTANTE", 28),
        ("J", "CARGO", 22), ("K", "NIVEL", 8), ("L", "LUGAR", 16),
        ("M", "Descripción de la Condición o Acto", 35), ("N", "Acción a implementar", 35),
        ("O", "Tipo de Causa", 25), ("P", "Controles Criticos", 18),
        ("Q", "Riesgo\n(A,M, B)", 12), ("R", "DESCRIPCION DE MEDIDA\n(EJECUTADO, EN PROCESO o PENDIENTE)", 22),
        ("S", "EMPRESA REPORTANTE", 18), ("T", "EMPRESA RESPONSABLE", 18), ("U", "PROCESO DE AVANCE", 18),
    ]
    for col_letter, name, width in headers:
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(row=5, column=col_idx, value=name)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = hdr_align; cell.border = thin_border
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = "A5:U5"
    meses = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
             "JULIO","AGOSTO","SETIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
    for idx, r in enumerate(reports, 1):
        row_num = 5 + idx
        dt_local = r.created_at
        fecha_reporte = r.fecha_reporte or dt_local
        mes = meses[dt_local.month - 1]
        semana = f"SEM {dt_local.isocalendar()[1]:02d}"
        guardia = wg_map.get((r.worker_name, r.group_name), "")
        cargo = cargo_map.get((r.worker_name, r.group_name), "")
        values = [
            idx, semana, mes, fecha_reporte.strftime("%d/%m/%Y"), "MINA",
            r.tipo, guardia, r.turno or "DÍA",
            r.worker_name, cargo, r.nivel or "",
            f"{r.ubicacion or ''} / Ref: {r.referencia or ''}",
            r.descripcion or "", r.accion_correctiva or "", r.tipo_descripcion or "",
            "", r.riesgo or "", "EJECUTADO",
            "Inversiones Prosol", "", "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = cell_font; cell.alignment = cell_align; cell.border = thin_border
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="BD_Acto_Condicion_Subestandar_{dt_mod.datetime.now().strftime("%Y%m%d")}.xlsx"'},
    )
