import os
import uuid
import io
import csv
from io import BytesIO
import shutil
import openpyxl
from datetime import date, time, datetime, timedelta
from typing import List, Optional
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Query, Request, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, StreamingResponse, Response
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session, joinedload, contains_eager

from sqlalchemy import func as _sf, text as _st

from .database import engine, get_db, Base, SessionLocal
from .excel_export import build_excel
from .daily_report import generate_daily_report, generate_daily_report_bytes, list_daily_reports
from .models import Worker, Report, JobEntry, JobImage, RacsReport, Guardia, WorkerGuardia, RacsWorker
from .models import EquipmentCategory, EquipmentSubitem, Collaborator, Macroprocess, WorkType, Action, Nivel, Turno
from .schemas import (
    ReportCreate, ReportResponse, WorkerResponse, JobImageResponse,
    RacsReportCreate, RacsReportResponse, RacsWorkerCreate, RacsWorkerUpdate,
)
from .config import settings, BASE_DIR

Base.metadata.create_all(bind=engine)

# ─── Seed config from hardcoded data (only if tables empty) ──────────
def _seed_hardcoded():
    db = SessionLocal()
    try:
        if db.query(Turno).count() > 0:
            return
        for i, t in enumerate(["Día", "Noche"]):
            db.add(Turno(name=t, position=i))
        for i, n in enumerate(["Nivel 05","Nivel 04","Nivel 03","Nivel 02","Nivel 01","Nivel 0","Nivel -01","Nivel -02","Nivel -03","Nivel -04","Nivel -05","Campamento","PTAR"]):
            db.add(Nivel(name=n, position=i))
        for i, (gk, n) in enumerate([
            ("Trackless", "Mantenimiento Mecánico"), ("Trackless", "Mantenimiento Eléctrico"),
            ("Convencional", "Mantenimiento Mecánico"), ("Convencional", "Mantenimiento Eléctrico"), ("Convencional", "Fabricaciones"), ("Convencional", "Instalaciones"),
            ("Electrico", "Mantenimiento Mecánico"), ("Electrico", "Mantenimiento Eléctrico"), ("Electrico", "Fabricaciones"), ("Electrico", "Instalaciones"),
        ]):
            db.add(Macroprocess(group_key=gk, name=n, position=i))
        for i, (tk, n) in enumerate([
            ("Trackless_Mecánico", "Auxilio Mecánico"), ("Trackless_Mecánico", "Bombas y Pistones"), ("Trackless_Mecánico", "Inspección y Correctivos Generales"), ("Trackless_Mecánico", "Inspección y Seguimiento en Labor"), ("Trackless_Mecánico", "Mantenimiento Preventivo PM 1"), ("Trackless_Mecánico", "Mantenimiento Preventivo PM 2"), ("Trackless_Mecánico", "Mantenimiento Preventivo PM 3"), ("Trackless_Mecánico", "Mantenimiento Preventivo PM 4"), ("Trackless_Mecánico", "Neumaticos"), ("Trackless_Mecánico", "Mantenimiento de Perforadora HC50 de Jumbo"), ("Trackless_Mecánico", "Sistemas de Dirección"), ("Trackless_Mecánico", "Sistema de Embrague"), ("Trackless_Mecánico", "Sistema de Frenos"), ("Trackless_Mecánico", "Sistemas Hidraulicos"), ("Trackless_Mecánico", "Sistemas de Suspensión"), ("Trackless_Mecánico", "Trabajos de Metalmecanica"),
            ("Trackless_Eléctrico", "Instalaciones Eléctricas, Iluminacion y Alarmas - Trackless"), ("Trackless_Eléctrico", "Inspecciones Generales y Seguimiento - Trackless"), ("Trackless_Eléctrico", "Mantenimiento Correctivo de Luces - Trackless"),
            ("Convencional_Mecánico", "Inspección y Seguimiento de equipo - General"), ("Convencional_Mecánico", "Correctivos Metalmecánica - Trackless"), ("Convencional_Mecánico", "Correcciones Metalmecánica - Convencional"), ("Convencional_Mecánico", "Mantenimiento Preventivo PM 1"), ("Convencional_Mecánico", "Mantenimiento Preventivo PM 2"), ("Convencional_Mecánico", "Mantenimiento Preventivo PM 3"), ("Convencional_Mecánico", "Mantenimiento Preventivo PM 4"),
            ("Convencional_Eléctrico", "Instalaciones Eléctricas - General"), ("Convencional_Eléctrico", "Mantenimiento Eléctrico - General"), ("Convencional_Eléctrico", "Tendido y Traslado de cables - General"), ("Convencional_Eléctrico", "Mantenimiento Preventivo Eléctrico - Convencional"), ("Convencional_Eléctrico", "Mantenimiento Correctivo Eléctrico - Convencional"), ("Convencional_Eléctrico", "Rebobinado de Motor"), ("Convencional_Eléctrico", "Telefonos"),
            ("Fabricacion_Soldadura", "Fabricaciones en General para Mina"), ("Fabricacion_Soldadura", "Fabricaciones en General para Taller"), ("Fabricacion_Soldadura", "Fabricaciones en General"), ("Fabricacion_Soldadura", "Fabricación de Alcayatas"), ("Fabricacion_Soldadura", "Correcciones de Fabricación - General"), ("Fabricacion_Soldadura", "Mantenimiento Preventivos - Convencional"),
            ("Instalaciones", "Instalaciones de Estructuras en General"), ("Instalaciones", "Instalaciones de Estructuras en Int. Mina"), ("Instalaciones", "Instalaciones de Estructuras en Campamento"), ("Instalaciones", "Telefonos"),
        ]):
            db.add(WorkType(type_key=tk, name=n, position=i))
        for i, (gk, n) in enumerate([
            ("Trackless", "Mantenimiento Preventivo"), ("Trackless", "Mantenimiento Correctivo"), ("Trackless", "Inspección (Liberación Eq.)"),
            ("Convencional_Electrico", "Mantenimiento Preventivo"), ("Convencional_Electrico", "Mantenimiento Correctivo"), ("Convencional_Electrico", "Instalación"), ("Convencional_Electrico", "Fabricación"), ("Convencional_Electrico", "Inspección (Liberación Eq.)"),
        ]):
            db.add(Action(group_key=gk, name=n, position=i))
        colab_data = {
            "Trackless": ["Wilson Apaza","Elvis Avalos","Yoel Castillo","Ronaldo Ccompara","Yordan Elguera","Miguel Espiritu","Jordi Gallegos","Jimy Huaman","Rivaldo Mamani","Ademer Moreto","Eddy Quispe","Felix Quispe","Jesús Quispe","Jesús Ramirez","Edwin Torres","Martin Sillocca"],
            "Convencional": ["Moises de la Cadena","Edgardo del Carpio","Erik Inca","Alan Llanquecha","Diego Machaca","Vlady mamani","Artemio Payehuanca","Fredy Taya","Edgar Zela","Aldair Vilca","Abraham Dionisio"],
            "Electrico": ["Ronaldo Ccompara","Jhonatan Chipane","Koos Coaquira","Wilder Espinal","Markc Huachaca","Jose Huayra","Santiago Huayra","Julio Perez","Wilber Quijahuaman","Rronar Surco"],
        }
        for gk, names in colab_data.items():
            for i, n in enumerate(names):
                db.add(Collaborator(group_name=gk, name=n, position=i))
        equipos_data = [
            ("Alimak", "Trackless", [("Alimak", "fin")]),
            ("Bomba de Pistón", "Trackless", [(f"Electrobomba de Pistón (EB.{i}-P)", "fin") for i in range(10, 17)]),
            ("Camionetas", "Convencional", [("Ambulancia (EUG-360)","kilometraje"),("Camioneta Conejo (VCP-772)","kilometraje"),("Camioneta GER-CORP (VDX-833)","kilometraje"),("Camioneta Gerencia (VDN-931)","kilometraje"),("Camioneta Gerencia (BZM-789)","kilometraje"),("Camioneta Lechera (VEG-809)","kilometraje"),("Camioneta Liebre 2 (VCW-854)","kilometraje"),("Camioneta Liebre 3 (VEE-949)","kilometraje"),("Camioneta Liebre 4 (VEH-817)","kilometraje"),("Camioneta Zorro (VAN-835)","kilometraje"),("Camioneta (VFF-923)","kilometraje"),("Camioneta (ERT-673)","kilometraje")]),
            ("Camioncitos", "Convencional", [("Camión de Azul (VCL-723)","kilometraje"),("Camión Tortuga 1 (VAP-905)","kilometraje"),("Camión Tortuga 2 (VDT-742)","kilometraje"),("Camión Tortuga 3 (VDT-726)","kilometraje"),("Camión Tortuga 4 (VEA-766)","kilometraje"),("Camión Tortuga 5 (VEE-770)","kilometraje"),("Camión Recuperación (VET-885)","kilometraje"),("Camión Rojo (VBR-892)","kilometraje")]),
            ("Cisternas", "Convencional", [("Cisterna de Agua (VDR-883)", "horometro_volquetes,kilometraje")]),
            ("Cargador Frontal", "Convencional", [("Cargador Frontal CAT 966 (CF-01)","horometro_motor"),("Cargador Frontal Liugong 856H","horometro_motor")]),
            ("Compresoras", "Convencional", [("Atlas Copco COE (GA 55)","horometro_motor"),("Atlas Copco (GA 90+) (RP 220)","horometro_motor"),("Atlas Copco (GA 90+) (Asunción)","horometro_motor"),("Atlas Copco (GA 110+)","horometro_motor"),("Atlas Copco (GA 110 VSD+)","horometro_motor"),("Atlas Copco (GA 160 VSD)","horometro_motor"),("Atlas Copco (GA 160 VSD+)","horometro_motor"),("Kaeser (SFC 200) N° 01","horometro_motor"),("Kaeser (SFC 220 N°) 02","horometro_motor")]),
            ("Dumper", "Trackless", [(f"Dumper N° {i:02d}", "fin") for i in range(1, 9)] + [("Dumper N° 10", "fin")]),
            ("Excavadora (CAT336)", "Convencional", [("Excavadora (CAT336)", "horometro_motor")]),
            ("Fabricaciones", "Convencional", [("Fabricaciones", "fin")]),
            ("Grupo Electrógeno", "Convencional", [("(GE-01) Asunción QAS 550","horometro_motor"),("(GE-02) Asunción QAS 550","horometro_motor"),("(GE-03) Asunción QAS 550","horometro_motor"),("(GE-04) Asunción MV 560","horometro_motor"),("(GE-05) Asunción QAS 500","horometro_motor"),("(GE-06) Asunción QAS 85","horometro_motor"),("(GE-07) Nivel 0 MV 560","horometro_motor"),("(GE-08) Rampa 220 MV 560","horometro_motor")]),
            ("Jumbos", "Trackless", [("Jumbo Komatsu ZJ21 (JM-01)","horometro_jumbo"),("Jumbo Muki N° 01 (JM-02)","horometro_jumbo"),("Jumbo Muki N° 02 (JM-03)","horometro_jumbo")]),
            ("Locomotoras", "Trackless", [(f"Locomotora N° {i:02d} (LOC-{i:02d})", "fin") for i in range(1, 11)]),
            ("Scooptrams", "Trackless", [("Scoop Komatsu WX07 (WX-07)","horometro_motor"),("Scoop Epiroc ST2G (ST2G)","horometro_motor"),("Scoop XCMG 1.5 YD (SD-03)","horometro_motor"),("Scoop XCMG 2.6 YD (SD-04)","horometro_motor")]),
            ("Retroexcavadora (REX-01)", "Convencional", [("Retroexcavadora (REX-01)", "horometro_motor")]),
            ("Palas Neumáticas", "Trackless", [(f"Pala N° {i:02d} (PALA-{i:02d})", "fin") for i in range(1, 11)]),
            ("Eq. Diamantina", "Trackless", [("Diamantina (PACKSAC)", "fin")]),
            ("Plataforma", "Trackless", [("Plataforma", "fin")]),
            ("Tableros", "Convencional", [("Tableros", "fin")]),
            ("Tendido de Cable", "Convencional", [("Tendido de Cable", "fin")]),
            ("Ventilador Eléctrico", "Trackless", [(f"Ventilador {'Centrífugo' if i%2 else 'Axial'} (V.{i:02d}-P)", "fin") for i in range(1, 35)]),
            ("Volquetes 4 Tn", "Convencional", [("Volquete 1 (VCI-946)","kilometraje"),("Volquete 2 (VCJ-739)","kilometraje"),("Volquete 3 (VDB-833)","kilometraje"),("Volquete 4 (VDJ-905)","kilometraje"),("Volquete 5 (VDM-742)","kilometraje")]),
            ("Volquetes - FMX", "Convencional", [("Volquete FMX 1 (BNK-709)","horometro_volquetes,kilometraje"),("Volquete FMX 2 (BDE-738)","horometro_volquetes,kilometraje"),("Volquete FMX 3 (VDJ-869)","horometro_volquetes,kilometraje"),("Volquete FMX 4 (VDJ-906)","horometro_volquetes,kilometraje"),("Volquete FMX 5 (VDF-703)","horometro_volquetes,kilometraje"),("Volquete FMX 6 (VDF-791)","horometro_volquetes,kilometraje"),("Volquete FMX 7 (CJB-903)","horometro_volquetes,kilometraje"),("Volquete FMX 8 (CJB-930)","horometro_volquetes,kilometraje"),("Volquete Shacman (CAR-839)","horometro_volquetes,kilometraje"),("Volquete Shacman (CAS-772)","horometro_volquetes,kilometraje")]),
            ("Winches Eléctricos", "Trackless", [(f"Winche Eléctrico N°{i:02d} (WE.{i:02d}-P)", "fin") for i in range(1, 5)]),
        ]
        for pos, (cat_name, act_grp, subs) in enumerate(equipos_data):
            cat = EquipmentCategory(name=cat_name, action_group=act_grp, position=pos)
            db.add(cat)
            db.flush()
            for spos, (sname, smeters) in enumerate(subs):
                db.add(EquipmentSubitem(category_id=cat.id, name=sname, meters=smeters, position=spos))
        db.commit()
        print("[OK] Hardcoded seed applied")
    except Exception as e:
        db.rollback()
        print(f"[ERROR] Hardcoded seed: {e}")
    finally:
        db.close()

# ─── Apply config_seed.json (only if tables empty) ────────────────────
import json as _json
_CONFIG_SEED = os.path.join(BASE_DIR, "config_seed.json")

def _apply_seed_from_json():
    if not os.path.exists(_CONFIG_SEED):
        return
    try:
        _sd = SessionLocal()
        if _sd.query(Turno).count() > 0:
            _sd.close()
            return
        with open(_CONFIG_SEED, "r", encoding="utf-8") as _f:
            _s = _json.load(_f)

        for i, t in enumerate(_s.get("turnos", [])):
            _sd.add(Turno(name=t["name"], position=i))
        for i, n in enumerate(_s.get("niveles", [])):
            _sd.add(Nivel(name=n["name"], position=i))
        for c in _s.get("colaboradores", []):
            _sd.add(Collaborator(group_name=c["group_name"], name=c["name"], position=c.get("position", 0)))
        for m in _s.get("macroprocesos", []):
            _sd.add(Macroprocess(group_key=m["group_key"], name=m["name"], position=m.get("position", 0)))
        for w in _s.get("tipos_trabajo", []):
            _sd.add(WorkType(type_key=w["type_key"], name=w["name"], default_action=w.get("default_action", ""), position=w.get("position", 0)))
        for a in _s.get("acciones", []):
            _sd.add(Action(group_key=a["group_key"], name=a["name"], position=a.get("position", 0)))
        for eq in _s.get("equipos", []):
            _cat = EquipmentCategory(name=eq["name"], action_group=eq.get("action_group", ""), position=eq.get("position", 0))
            _sd.add(_cat)
            _sd.flush()
            for j, sub in enumerate(eq.get("subitems", [])):
                _sd.add(EquipmentSubitem(category_id=_cat.id, name=sub["name"], meters=sub.get("meters", "fin"), position=j))
        _sd.commit()
        total = sum(len(v) for v in _s.values())
        print(f"[OK] Seed applied from config_seed.json ({total} records)")
    except Exception as _e:
        _sd.rollback()
        print(f"[WARN] Could not apply config_seed.json: {_e}")
    finally:
        _sd.close()

# JSON seed first (primary), hardcoded seed second (fallback if no JSON)
_apply_seed_from_json()
_seed_hardcoded()

# ─── Ensure schema columns exist ─────────────────────────────────────────
_try_alter = lambda sql, msg: (_d := SessionLocal(), _d.execute(_st(sql)), _d.commit(), _d.close(), print(msg)) if None else None
try:
    _d = SessionLocal(); _d.execute(_st('ALTER TABLE work_types ADD COLUMN default_action VARCHAR(200)')); _d.commit(); _d.close(); print("[MIGRACION] Columna default_action agregada a work_types")
except: _d.rollback(); _d.close()
try:
    _d = SessionLocal(); _d.execute(_st('ALTER TABLE job_entries ADD COLUMN collaborators TEXT')); _d.commit(); _d.close(); print("[MIGRACION] Columna collaborators agregada a job_entries")
except: _d.rollback(); _d.close()
try:
    _d = SessionLocal()
    for col in ['riesgo', 'accion_correctiva', 'tipo_descripcion']:
        try:
            _d.execute(_st(f'ALTER TABLE racs_reports ADD COLUMN {col} TEXT'))
            _d.commit()
            print(f"[MIGRACION] Columna {col} agregada a racs_reports")
        except:
            _d.rollback()
    _d.close()
except: pass
try:
    _d = SessionLocal()
    _d.execute(_st("ALTER TABLE worker_guardias ADD COLUMN cargo VARCHAR(200)"))
    _d.commit()
    _d.close()
    print("[MIGRACION] Columna cargo agregada a worker_guardias")
except: _d.rollback(); _d.close()

# ─── Create indexes ─────────────────────────────────────────────────────
for _sql in [
    "CREATE INDEX IF NOT EXISTS ix_reports_date_id ON reports(date, id)",
    "CREATE INDEX IF NOT EXISTS ix_reports_group_shift_date ON reports(group_name, shift, date)",
    "CREATE INDEX IF NOT EXISTS ix_reports_created_at_id ON reports(created_at, id)",
    "CREATE INDEX IF NOT EXISTS ix_job_entries_report_id ON job_entries(report_id)",
]:
    try:
        _d = SessionLocal(); _d.execute(_st(_sql)); _d.commit(); _d.close()
    except: _d.rollback(); _d.close()
try:
    _d = SessionLocal()
    for col in ['turno', 'categoria']:
        try:
            _d.execute(_st(f'ALTER TABLE racs_reports ADD COLUMN {col} VARCHAR(50)'))
            _d.commit()
            print(f"[MIGRACION] Columna {col} agregada a racs_reports")
        except:
            _d.rollback()
    _d.close()
except: pass
try:
    _d = SessionLocal()
    _d.execute(_st("ALTER TABLE racs_reports ADD COLUMN referencia VARCHAR(300)"))
    _d.commit()
    _d.close()
    print("[MIGRACION] Columna referencia agregada a racs_reports")
except: _d.rollback(); _d.close()
try:
    _d = SessionLocal()
    _d.execute(_st("ALTER TABLE racs_reports ADD COLUMN nivel VARCHAR(100)"))
    _d.commit()
    _d.close()
    print("[MIGRACION] Columna nivel agregada a racs_reports")
except: _d.rollback(); _d.close()
try:
    _d = SessionLocal()
    _d.execute(_st("ALTER TABLE racs_reports ADD COLUMN fecha_reporte DATE"))
    _d.commit()
    _d.close()
    print("[MIGRACION] Columna fecha_reporte agregada a racs_reports")
except: _d.rollback(); _d.close()
print("[OK] Indices verificados")

ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(os.path.dirname(BASE_DIR), "static")
UPLOAD_DIR = os.path.join(STATIC_DIR, "uploads")
TEMPLATE_DIR = os.path.join(BASE_DIR, "app", "templates")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(TEMPLATE_DIR, exist_ok=True)

templates = Jinja2Templates(directory=TEMPLATE_DIR)

app = FastAPI(title="Reporte Diario de Mantenimiento", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/", response_class=HTMLResponse)
def index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


def parse_time(value):
    if not value:
        return None
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p"):
        try:
            parts = value.strip().split(":")
            if len(parts) == 2:
                return time(int(parts[0]), int(parts[1]))
            elif len(parts) == 3:
                return time(int(parts[0]), int(parts[1]), int(parts[2]))
        except (ValueError, IndexError):
            continue
    return None


def parse_date(value):
    if not value:
        return date.today()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue
    return date.today()


# ─── API Routes ───────────────────────────────────────────────────────────────

@app.get("/api/workers", response_model=List[WorkerResponse])
def list_workers(db: Session = Depends(get_db)):
    return db.query(Worker).order_by(Worker.name).all()


@app.post("/api/workers")
def create_worker(name: str = Form(...), email: str = Form(...), group_name: str = Form(""), db: Session = Depends(get_db)):
    worker = Worker(name=name, email=email, group_name=group_name)
    db.add(worker)
    db.commit()
    db.refresh(worker)
    return worker


@app.post("/api/reports", response_model=ReportResponse)
def create_report(report: ReportCreate, db: Session = Depends(get_db)):
    worker = db.query(Worker).filter(
        Worker.name == report.worker_name,
        Worker.email == report.worker_email
    ).first()
    if not worker:
        worker = Worker(name=report.worker_name, email=report.worker_email)
        db.add(worker)
        db.flush()

    # Populate report-level collaborator fields from entries (backward compat)
    all_colabs = {
        g: [] for g in ["Trackless", "Convencional", "Electrico"]
    }
    for entry_data in report.entries:
        if entry_data.collaborators:
            g_key = "Trackless" if "Trackless" in (report.group_name or "") else "Convencional" if "Convencional" in (report.group_name or "") else "Electrico"
            all_colabs[g_key].append(entry_data.collaborators)

    db_report = Report(
        worker_id=worker.id,
        date=parse_date(report.date),
        shift=report.shift,
        group_name=report.group_name,
        start_time=parse_time(report.start_time),
        end_time=parse_time(report.end_time),
        collaborators_trackless=", ".join(all_colabs["Trackless"]) or report.collaborators_trackless,
        collaborators_convencional=", ".join(all_colabs["Convencional"]) or report.collaborators_convencional,
        collaborators_electrico=", ".join(all_colabs["Electrico"]) or report.collaborators_electrico,
    )
    db.add(db_report)
    db.flush()

    for entry_data in report.entries:
        db_entry = JobEntry(
            report_id=db_report.id,
            macroprocess=entry_data.macroprocess,
            work_type=entry_data.work_type,
            work_subtype=entry_data.work_subtype,
            action=entry_data.action,
            description=entry_data.description,
            level=entry_data.level,
            location=entry_data.location,
            start_time_int=parse_time(entry_data.start_time_int),
            end_time_int=parse_time(entry_data.end_time_int),
            duration=entry_data.duration,
            equipment=entry_data.equipment,
            horometer_motor=entry_data.horometer_motor,
            horometer_motor_jumbo=entry_data.horometer_motor_jumbo,
            horometer_motor_volquetes=entry_data.horometer_motor_volquetes,
            horometer_electric=entry_data.horometer_electric,
            horometer_percussion=entry_data.horometer_percussion,
            kilometer=entry_data.kilometer,
            blanco=entry_data.blanco,
            collaborators=entry_data.collaborators,
        )
        db.add(db_entry)
    db.commit()
    db.refresh(db_report)
    return db_report


@app.get("/api/reports")
def list_reports(
    cursor_id: int = Query(None, alias="cursor"),
    cursor_date: str = Query(None, alias="cursor_date"),
    limit: int = Query(50),
    worker_name: str = Query(None),
    date_from: str = Query(None),
    date_to: str = Query(None),
    db: Session = Depends(get_db),
):
    try:
        q = db.query(Report).options(joinedload(Report.worker))
        if worker_name:
            q = q.join(Worker).filter(Worker.name.ilike(f"%{worker_name}%"))
        if date_from:
            q = q.filter(Report.date >= parse_date(date_from))
        if date_to:
            q = q.filter(Report.date <= parse_date(date_to))
        # Keyset pagination: (created_at, id) < (cursor_date, cursor_id)
        if cursor_id is not None and cursor_date is not None:
            try:
                dt = datetime.fromisoformat(cursor_date)
            except Exception:
                dt = None
            if dt is not None:
                q = q.filter(Report.created_at < dt).filter(Report.id < cursor_id)
        reports = q.order_by(Report.created_at.desc(), Report.id.desc()).limit(limit + 1).all()
        has_more = len(reports) > limit
        reports = reports[:limit]
        result = []
        for rp in reports:
            d = {
                "id": rp.id,
                "worker_id": rp.worker_id,
                "worker_name": rp.worker.name if rp.worker else f"ID {rp.worker_id}",
                "date": rp.date,
                "shift": rp.shift,
                "group_name": rp.group_name,
                "start_time": rp.start_time,
                "end_time": rp.end_time,
                "collaborators_trackless": rp.collaborators_trackless,
                "collaborators_convencional": rp.collaborators_convencional,
                "collaborators_electrico": rp.collaborators_electrico,
                "created_at": rp.created_at.isoformat() if rp.created_at else None,
                "entries": [{
                    "id": e.id,
                    "report_id": e.report_id,
                    "macroprocess": e.macroprocess,
                    "work_type": e.work_type,
                    "work_subtype": e.work_subtype,
                    "action": e.action,
                    "description": e.description,
                    "level": e.level,
                    "location": e.location,
                    "start_time_int": e.start_time_int.strftime("%H:%M") if e.start_time_int else None,
                    "end_time_int": e.end_time_int.strftime("%H:%M") if e.end_time_int else None,
                    "duration": e.duration,
                    "equipment": e.equipment,
                    "horometer_motor": e.horometer_motor,
                    "horometer_motor_jumbo": e.horometer_motor_jumbo,
                    "horometer_motor_volquetes": e.horometer_motor_volquetes,
                    "horometer_electric": e.horometer_electric,
                    "horometer_percussion": e.horometer_percussion,
                    "kilometer": e.kilometer,
                    "blanco": e.blanco,
                    "collaborators": e.collaborators,
                    "images": [{"id": img.id, "filename": img.filename, "original_name": img.original_name} for img in e.images],
                } for e in rp.entries],
            }
            result.append(d)
        last_item = reports[-1] if reports else None
        return {
            "data": result,
            "next_cursor": (last_item.id, last_item.created_at.isoformat()) if last_item and has_more else None,
            "has_more": has_more,
        }
    except Exception as exc:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Error al listar reportes: {exc}")


@app.get("/api/reports/export/excel")
def export_reports_excel(
    date_from: str = Query(None),
    date_to: str = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Report).options(joinedload(Report.worker)).options(joinedload(Report.entries).joinedload(JobEntry.images))
    if date_from:
        q = q.filter(Report.date >= parse_date(date_from))
    if date_to:
        q = q.filter(Report.date <= parse_date(date_to))
    reports = q.order_by(Report.created_at.asc()).all()

    data = []
    for rp in reports:
        d = {
            "id": rp.id,
            "worker_id": rp.worker_id,
            "worker_name": rp.worker.name if rp.worker else None,
            "date": rp.date,
            "shift": rp.shift,
            "group_name": rp.group_name,
            "start_time": rp.start_time,
            "end_time": rp.end_time,
            "collaborators_trackless": rp.collaborators_trackless,
            "collaborators_convencional": rp.collaborators_convencional,
            "collaborators_electrico": rp.collaborators_electrico,
            "created_at": rp.created_at,
            "entries": [{
                "id": e.id,
                "macroprocess": e.macroprocess,
                "work_type": e.work_type,
                "work_subtype": e.work_subtype,
                "action": e.action,
                "description": e.description,
                "level": e.level,
                "location": e.location,
                "start_time_int": e.start_time_int,
                "end_time_int": e.end_time_int,
                "duration": e.duration,
                "equipment": e.equipment,
                "horometer_motor": e.horometer_motor,
                "horometer_motor_jumbo": e.horometer_motor_jumbo,
                "horometer_motor_volquetes": e.horometer_motor_volquetes,
                "horometer_electric": e.horometer_electric,
                "horometer_percussion": e.horometer_percussion,
                "kilometer": e.kilometer,
                "blanco": e.blanco,
                "collaborators": e.collaborators,
            } for e in rp.entries],
        }
        data.append(d)

    wb = build_excel(data)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_mantenimiento.xlsx"},
    )


@app.get("/api/reports/export/csv")
def export_reports_csv(
    date_from: str = Query(None),
    date_to: str = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Report).join(Worker)
    if date_from:
        q = q.filter(Report.date >= parse_date(date_from))
    if date_to:
        q = q.filter(Report.date <= parse_date(date_to))
    reports = q.order_by(Report.created_at.desc()).all()

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["ID Reporte","Fecha","Turno","Grupo","Colaboradores del Trabajo",
                "Trabajo #","Macroproceso","Tipo Trabajo","Acción","Descripción","Equipo","Nivel","Lugar",
                "Inicio","Fin","Duración","Hor. Motor","Hor. Jumbo","Hor. Volquetes","Hor. Eléctrico","Hor. Percusión","Kilometraje"])

    for rp in reports:
        entries = rp.entries if hasattr(rp, 'entries') and rp.entries else [None]
        for entry in entries:
            colabs = entry.collaborators if entry and entry.collaborators else ""
            w.writerow([
                rp.id, rp.date, rp.shift, rp.group_name,
                colabs,
                entry.id if entry else "", entry.macroprocess if entry else "", entry.work_type if entry else "",
                entry.action if entry else "", entry.description if entry else "", entry.equipment if entry else "",
                entry.level if entry else "", entry.location if entry else "",
                entry.start_time_int if entry else "", entry.end_time_int if entry else "", entry.duration if entry else "",
                entry.horometer_motor if entry else "", entry.horometer_motor_jumbo if entry else "",
                entry.horometer_motor_volquetes if entry else "", entry.horometer_electric if entry else "",
                entry.horometer_percussion if entry else "", entry.kilometer if entry else "",
            ])

    output.seek(0)
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=reportes.csv"})


# ─── Daily Report Endpoints ─────────────────────────────────────────────────

@app.get("/api/daily-reports")
def get_daily_reports(db: Session = Depends(get_db)):
    """List dates with available reports (from DB)."""
    from sqlalchemy import func as _sf
    rows = db.query(Report.date).distinct().order_by(Report.date.desc()).all()
    return [{"date": r[0].isoformat() if hasattr(r[0], 'isoformat') else str(r[0])} for r in rows]


@app.post("/api/daily-reports/generate")
def generate_daily(target_date: str = Body("", embed=True), db: Session = Depends(get_db)):
    """Generate daily report for a given date (YYYY-MM-DD). Defaults to yesterday."""
    if target_date and target_date.strip():
        d = parse_date(target_date.strip())
    else:
        d = (datetime.now() - timedelta(days=1)).date()

    if d is None:
        raise HTTPException(400, "Fecha invalida")

    # Fetch all reports for this date, grouped by group_name and shift
    reports = db.query(Report).options(
        joinedload(Report.worker), joinedload(Report.entries)
    ).filter(Report.date == d).all()

    if not reports:
        raise HTTPException(404, f"No hay reportes para {d.isoformat()}")

    # Group data by group_name → shift
    grouped = {}
    for rp in reports:
        group = rp.group_name
        shift = rp.shift
        grouped.setdefault(group, {}).setdefault(shift, []).append({
            "id": rp.id,
            "date": rp.date,
            "shift": shift,
            "group_name": group,
            "collaborators_trackless": rp.collaborators_trackless,
            "collaborators_convencional": rp.collaborators_convencional,
            "collaborators_electrico": rp.collaborators_electrico,
            "entries": [{
                "macroprocess": e.macroprocess,
                "work_type": e.work_type,
                "work_subtype": e.work_subtype,
                "action": e.action,
                "description": e.description,
                "level": e.level,
                "location": e.location,
                "start_time_int": e.start_time_int,
                "end_time_int": e.end_time_int,
                "duration": e.duration,
                "equipment": e.equipment,
                "horometer_motor": e.horometer_motor,
                "horometer_motor_jumbo": e.horometer_motor_jumbo,
                "horometer_motor_volquetes": e.horometer_motor_volquetes,
                "horometer_electric": e.horometer_electric,
                "horometer_percussion": e.horometer_percussion,
                "kilometer": e.kilometer,
                "collaborators": e.collaborators,
            } for e in rp.entries],
        })

    return _stream_report(d, grouped)


def _stream_report(d: date, grouped: dict):
    """Generate report in memory and return as file download response."""
    from fastapi.responses import StreamingResponse
    buf = generate_daily_report_bytes(d, grouped)
    fname = f"Reporte_Diario_{d.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@app.get("/api/reports/{report_id}")
def get_report(report_id: int, db: Session = Depends(get_db)):
    report = db.query(Report).options(joinedload(Report.worker)).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(404, "Reporte no encontrado")
    d = {
        "id": report.id,
        "worker_id": report.worker_id,
        "worker_name": report.worker.name if report.worker else f"ID {report.worker_id}",
        "date": report.date,
        "shift": report.shift,
        "group_name": report.group_name,
        "start_time": report.start_time,
        "end_time": report.end_time,
        "collaborators_trackless": report.collaborators_trackless,
        "collaborators_convencional": report.collaborators_convencional,
        "collaborators_electrico": report.collaborators_electrico,
        "created_at": report.created_at.isoformat() if report.created_at else None,
        "entries": [{
            "id": e.id,
            "report_id": e.report_id,
            "macroprocess": e.macroprocess,
            "work_type": e.work_type,
            "work_subtype": e.work_subtype,
            "action": e.action,
            "description": e.description,
            "level": e.level,
            "location": e.location,
            "start_time_int": e.start_time_int.strftime("%H:%M") if e.start_time_int else None,
            "end_time_int": e.end_time_int.strftime("%H:%M") if e.end_time_int else None,
            "duration": e.duration,
            "equipment": e.equipment,
            "horometer_motor": e.horometer_motor,
            "horometer_motor_jumbo": e.horometer_motor_jumbo,
            "horometer_motor_volquetes": e.horometer_motor_volquetes,
            "horometer_electric": e.horometer_electric,
            "horometer_percussion": e.horometer_percussion,
            "kilometer": e.kilometer,
            "blanco": e.blanco,
            "collaborators": e.collaborators,
            "images": [{"id": img.id, "filename": img.filename, "original_name": img.original_name} for img in e.images],
        } for e in report.entries],
    }
    return d


DELETE_REPORT_PASSWORD = "70212352"


@app.delete("/api/reports/{report_id}")
def delete_report(report_id: int, password: str = Query(...), db: Session = Depends(get_db)):
    if password != DELETE_REPORT_PASSWORD:
        raise HTTPException(401, "Contraseña incorrecta")
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(404, "Reporte no encontrado")
    db.delete(report)
    db.commit()
    return {"ok": True, "deleted_id": report_id}


@app.post("/api/reports/batch-delete")
def batch_delete_reports(ids: list[int] = Body(...), password: str = Body(...), db: Session = Depends(get_db)):
    if password != DELETE_REPORT_PASSWORD:
        raise HTTPException(401, "Contraseña incorrecta")
    deleted = []
    for rid in ids:
        report = db.query(Report).filter(Report.id == rid).first()
        if report:
            db.delete(report)
            deleted.append(rid)
    db.commit()
    return {"ok": True, "deleted_ids": deleted}


@app.post("/api/reports/delete-verify")
def verify_delete_password(password: str = Body(..., embed=True)):
    if password == DELETE_REPORT_PASSWORD:
        return {"ok": True}
    raise HTTPException(401, "Contraseña incorrecta")


@app.put("/api/reports/{report_id}")
def update_report(report_id: int, data: dict = Body(...), db: Session = Depends(get_db)):
    if data.get("password") != DELETE_REPORT_PASSWORD:
        raise HTTPException(401, "Contraseña incorrecta")
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise HTTPException(404, "Reporte no encontrado")
    if "date" in data:
        report.date = parse_date(data["date"])
    if "shift" in data:
        report.shift = data["shift"]
    if "group_name" in data:
        report.group_name = data["group_name"]
    db.commit()
    db.refresh(report)
    return {"ok": True, "id": report.id}


@app.post("/api/reports/{entry_id}/images")
async def upload_image(entry_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    entry = db.query(JobEntry).filter(JobEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Entrada de trabajo no encontrada")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Formato no permitido: {ext}. Use: jpg, png, gif, webp")

    content = await file.read()
    if len(content) > settings.max_image_size_mb * 1024 * 1024:
        raise HTTPException(400, f"Imagen muy grande. Máx: {settings.max_image_size_mb}MB")

    # Store in date-based folders to avoid 100k files in one dir
    today = datetime.now().strftime("%Y/%m/%d")
    rel_dir = today.replace("/", os.sep)
    dest_dir = os.path.join(UPLOAD_DIR, rel_dir)
    os.makedirs(dest_dir, exist_ok=True)

    unique_name = f"{uuid.uuid4().hex}{ext}"
    rel_path = os.path.join(rel_dir, unique_name)
    file_path = os.path.join(UPLOAD_DIR, rel_path)

    with open(file_path, "wb") as f:
        f.write(content)

    db_image = JobImage(
        entry_id=entry_id,
        filename=rel_path,
        original_name=file.filename,
    )
    db.add(db_image)
    db.commit()
    db.refresh(db_image)
    return db_image


@app.delete("/api/images/{image_id}")
def delete_image(image_id: int, db: Session = Depends(get_db)):
    image = db.query(JobImage).filter(JobImage.id == image_id).first()
    if not image:
        raise HTTPException(404, "Imagen no encontrada")

    file_path = os.path.join(UPLOAD_DIR, image.filename)
    if os.path.exists(file_path):
        os.remove(file_path)

    db.delete(image)
    db.commit()
    return {"ok": True}


@app.get("/api/equipment/{equipo}/last-reading")
def get_equipment_last_reading(equipo: str, db: Session = Depends(get_db)):
    """Return the last horometer/kilometer readings for a given piece of equipment."""
    entry = (
        db.query(JobEntry)
        .filter(JobEntry.equipment == equipo)
        .order_by(JobEntry.id.desc())
        .first()
    )
    if not entry:
        return {"found": False, "readings": {}}
    readings = {
        "horometer_motor": entry.horometer_motor,
        "horometer_motor_jumbo": entry.horometer_motor_jumbo,
        "horometer_motor_volquetes": entry.horometer_motor_volquetes,
        "horometer_electric": entry.horometer_electric,
        "horometer_percussion": entry.horometer_percussion,
        "kilometer": entry.kilometer,
    }
    return {"found": True, "readings": readings}


@app.get("/api/options")
def get_options(db: Session = Depends(get_db)):
    turnos = [t.name for t in db.query(Turno).order_by(Turno.position).all()]
    grupos = ["Mantt. Eq. Trackless", "Mantt. Eq. Convencional", "Mantt. Eq. Electrico"]

    colab = {}
    for c in db.query(Collaborator).order_by(Collaborator.position).all():
        colab.setdefault(c.group_name, []).append(c.name)

    macro = {}
    for m in db.query(Macroprocess).order_by(Macroprocess.position).all():
        macro.setdefault(m.group_key, []).append(m.name)

    wt = {}
    for w in db.query(WorkType).order_by(WorkType.position).all():
        wt.setdefault(w.type_key, []).append({"name": w.name, "default_action": w.default_action or ""})

    acc = {}
    for a in db.query(Action).order_by(Action.position).all():
        acc.setdefault(a.group_key, []).append(a.name)

    niveles = [n.name for n in db.query(Nivel).order_by(Nivel.position).all()]

    equipos = {}
    for cat in db.query(EquipmentCategory).order_by(EquipmentCategory.position).all():
        subs = []
        for sub in cat.subitems:
            subs.append({"nombre": sub.name, "mide": sub.meters})
        equipos[cat.name] = {"subequipos": subs, "accion_grupo": cat.action_group or ""}

    return {
        "turnos": turnos,
        "grupos": grupos,
        "colaboradores": colab,
        "colab_max_select": 4,
        "macroprocesos": macro,
        "tipos_trabajo": wt,
        "acciones": acc,
        "niveles": niveles,
        "equipos": equipos,
    }


# ─── Dashboard ────────────────────────────────────────────────────────────────

def _parse_duration_minutes(dur):
    if not dur:
        return None
    try:
        parts = dur.strip().split(":")
        if len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
        if len(parts) == 3:
            return int(parts[0]) * 60 + int(parts[1])
    except (ValueError, IndexError):
        pass
    return None


def _build_filtered_query(db, date_from, date_to, group):
    """Return a base query of JobEntry joined to Report with filters applied."""
    q = db.query(JobEntry).join(JobEntry.report).options(contains_eager(JobEntry.report))
    if date_from:
        q = q.filter(Report.date >= parse_date(date_from))
    if date_to:
        q = q.filter(Report.date <= parse_date(date_to))
    if group:
        q = q.filter(Report.group_name == group)
    return q


def _compute_period_stats(db, date_from, date_to, group):
    """Compute all stats for a single period."""
    q_base = _build_filtered_query(db, date_from, date_to, group)

    # Total jobs
    total_jobs = q_base.with_entities(_sf.count(JobEntry.id)).scalar() or 0

    # Distinct workers
    workers = (
        db.query(Worker.id, Worker.name, Worker.group_name)
        .join(Report)
    )
    if date_from:
        workers = workers.filter(Report.date >= parse_date(date_from))
    if date_to:
        workers = workers.filter(Report.date <= parse_date(date_to))
    if group:
        workers = workers.filter(Report.group_name == group)
    workers = workers.distinct().all()
    worker_count = len(workers)

    # Distinct collaborators
    collab_q = q_base.with_entities(JobEntry.collaborators).filter(JobEntry.collaborators.isnot(None))
    collab_names = set()
    for (c,) in collab_q.all():
        for name in c.split(","):
            name = name.strip()
            if name:
                collab_names.add(name)

    # Total duration hours
    dur_q = q_base.with_entities(JobEntry.duration).filter(JobEntry.duration.isnot(None), JobEntry.duration != "")
    total_minutes = 0
    for (d,) in dur_q.all():
        m = _parse_duration_minutes(d)
        if m is not None:
            total_minutes += m
    total_hours = round(total_minutes / 60, 1) if total_minutes else 0

    # Average job duration
    avg_minutes = round(total_minutes / total_jobs, 1) if total_jobs else 0

    # Jobs by equipment
    equip_data = (
        q_base.with_entities(JobEntry.equipment, JobEntry.duration)
        .filter(JobEntry.equipment.isnot(None), JobEntry.equipment != "")
        .all()
    )
    from collections import defaultdict
    equip_durs = defaultdict(list)
    for equip, dur in equip_data:
        if dur:
            equip_durs[equip].append(dur)
    jobs_by_equipment = []
    for equip, durs in sorted(equip_durs.items(), key=lambda x: -len(x[1])):
        cnt = len(durs)
        mins = [_parse_duration_minutes(d) for d in durs]
        valid = [m for m in mins if m is not None]
        avg = round(sum(valid) / len(valid), 1) if valid else None
        equip_hours = round((sum(valid) / 60), 1) if valid else 0
        jobs_by_equipment.append({
            "equipment": equip, "count": cnt,
            "avg_duration_min": avg, "total_hours": equip_hours,
        })

    # Jobs by team
    team_q = (
        q_base.with_entities(Report.group_name, _sf.count(JobEntry.id))
        .group_by(Report.group_name)
        .order_by(_sf.count(JobEntry.id).desc())
        .all()
    )
    jobs_by_team = [{"team": g, "count": c} for g, c in team_q]

    # Avg duration by team
    team_dur_q = (
        q_base.with_entities(Report.group_name, JobEntry.duration)
        .filter(JobEntry.duration.isnot(None), JobEntry.duration != "")
        .all()
    )
    team_durs = defaultdict(list)
    for team, dur in team_dur_q:
        if dur:
            team_durs[team].append(dur)
    avg_dur_by_team = []
    for team, durs in team_durs.items():
        mins = [_parse_duration_minutes(d) for d in durs]
        valid = [m for m in mins if m is not None]
        avg = round(sum(valid) / len(valid), 1) if valid else 0
        avg_dur_by_team.append({"team": team, "avg_min": avg})

    # Jobs per day (trend)
    day_q = (
        q_base.with_entities(Report.date, _sf.count(JobEntry.id))
        .group_by(Report.date)
        .order_by(Report.date)
        .all()
    )
    jobs_per_day = [{"date": str(d), "count": c} for d, c in day_q]

    # Shift comparison
    shift_q = (
        q_base.with_entities(Report.shift, _sf.count(JobEntry.id))
        .group_by(Report.shift)
        .all()
    )
    shift_comparison = {s or "Sin turno": c for s, c in shift_q}

    # PM vs CM ratio (action-based: Preventivo ≈ PM, Correctivo ≈ CM)
    action_q = (
        q_base.with_entities(JobEntry.action, _sf.count(JobEntry.id))
        .filter(JobEntry.action.isnot(None), JobEntry.action != "")
        .group_by(JobEntry.action)
        .all()
    )
    pm_count = sum(c for a, c in action_q if "Preventivo" in (a or ""))
    cm_count = sum(c for a, c in action_q if "Correctivo" in (a or ""))
    pm_cm_ratio = {
        "pm": pm_count,
        "cm": cm_count,
        "other": total_jobs - pm_count - cm_count,
        "pm_pct": round(pm_count / total_jobs * 100, 1) if total_jobs else 0,
        "cm_pct": round(cm_count / total_jobs * 100, 1) if total_jobs else 0,
    }

    # Macroprocess distribution
    macro_q = (
        q_base.with_entities(JobEntry.macroprocess, _sf.count(JobEntry.id))
        .filter(JobEntry.macroprocess.isnot(None), JobEntry.macroprocess != "")
        .group_by(JobEntry.macroprocess)
        .order_by(_sf.count(JobEntry.id).desc())
        .all()
    )
    macroprocess_dist = [{"macroprocess": m or "Sin especificar", "count": c} for m, c in macro_q]

    # Top collaborators
    collab_counter = {}
    for (c,) in collab_q.all():
        for name in c.split(","):
            n = name.strip()
            if n:
                collab_counter[n] = collab_counter.get(n, 0) + 1
    top_collaborators = [{"name": k, "count": v} for k, v in
                         sorted(collab_counter.items(), key=lambda x: -x[1])[:15]]

    # Weekly summaries (ISO weeks within period)
    weeks = {}
    for d_str, cnt in day_q:
        d = d_str if isinstance(d_str, date) else parse_date(str(d_str))
        if d is None:
            continue
        iso = d.isocalendar()
        week_key = f"{iso[0]}-W{iso[1]:02d}"
        if week_key not in weeks:
            # Find week start (Monday)
            monday = d - timedelta(days=d.weekday())
            weeks[week_key] = {"week": week_key, "start": str(monday), "count": 0, "by_team": {}}
        weeks[week_key]["count"] += cnt
    # Get team breakdown per week
    for week_key in weeks:
        ws = parse_date(weeks[week_key]["start"])
        we = ws + timedelta(days=6)
        wt_q = (
            q_base.with_entities(Report.group_name, _sf.count(JobEntry.id))
            .filter(Report.date >= ws, Report.date <= we)
            .group_by(Report.group_name)
            .all()
        )
        weeks[week_key]["by_team"] = {g or "Sin grupo": c for g, c in wt_q}
    weekly_summaries = sorted(weeks.values(), key=lambda x: x["week"])

    return {
        "kpis": {
            "total_jobs": total_jobs,
            "total_hours": total_hours,
            "worker_count": worker_count,
            "collaborator_count": len(collab_names),
            "avg_duration_min": avg_minutes,
            "pm_pct": pm_cm_ratio["pm_pct"],
            "cm_pct": pm_cm_ratio["cm_pct"],
        },
        "workers": [{"id": w.id, "name": w.name, "team": w.group_name} for w in workers],
        "jobs_by_equipment": jobs_by_equipment,
        "top_equipment": jobs_by_equipment[:10],
        "jobs_per_day": jobs_per_day,
        "jobs_by_team": jobs_by_team,
        "avg_duration_by_team": avg_dur_by_team,
        "shift_comparison": shift_comparison,
        "pm_cm_ratio": pm_cm_ratio,
        "macroprocess_dist": macroprocess_dist,
        "top_collaborators": top_collaborators,
        "weekly_summaries": weekly_summaries,
    }


def _compute_okrs(stats):
    """Generate OKR suggestions based on actual data."""
    pct = stats["pm_cm_ratio"]["pm_pct"]
    cm_pct = stats["pm_cm_ratio"]["cm_pct"]
    avg_min = stats["kpis"]["avg_duration_min"]
    total_jobs = stats["kpis"]["total_jobs"]
    total_hours = stats["kpis"]["total_hours"]
    workers = stats["kpis"]["worker_count"]
    collabs = stats["kpis"]["collaborator_count"]

    okrs = []

    # OKR 1: Maintenance quality
    okrs.append({
        "objective": "Mejorar la Calidad del Mantenimiento",
        "key_results": [
            {
                "kr": f"Aumentar mantenimiento preventivo de {pct}% a ≥70%",
                "current": f"{pct}%",
                "target": "≥70%",
                "progress": "on_track" if pct >= 50 else "needs_attention" if pct >= 30 else "critical",
            },
            {
                "kr": f"Reducir mantenimiento correctivo de {cm_pct}% a ≤30%",
                "current": f"{cm_pct}%",
                "target": "≤30%",
                "progress": "on_track" if cm_pct <= 30 else "needs_attention" if cm_pct <= 50 else "critical",
            },
            {
                "kr": f"Reducir tiempo promedio por trabajo de {avg_min} min a ≤60 min",
                "current": f"{avg_min} min",
                "target": "≤60 min",
                "progress": "on_track" if avg_min <= 60 else "needs_attention" if avg_min <= 90 else "critical",
            },
        ]
    })

    # OKR 2: Productivity
    jobs_per_worker = round(total_jobs / workers, 1) if workers else 0
    okrs.append({
        "objective": "Optimizar la Productividad del Equipo",
        "key_results": [
            {
                "kr": f"Mantener ≥4 trabajos por trabajador ({jobs_per_worker} actual)",
                "current": f"{jobs_per_worker}",
                "target": "≥4",
                "progress": "on_track" if jobs_per_worker >= 4 else "needs_attention" if jobs_per_worker >= 2.5 else "critical",
            },
            {
                "kr": f"Aumentar colaboradores activos de {collabs} a ≥{min(collabs + 5, 40)}",
                "current": f"{collabs}",
                "target": f"≥{min(collabs + 5, 40)}",
                "progress": "on_track" if collabs >= 15 else "needs_attention",
            },
        ]
    })

    # OKR 3: Data-driven
    okrs.append({
        "objective": "Fortalecer la Gestión Basada en Datos",
        "key_results": [
            {
                "kr": "Registrar 100% de trabajos en plataforma digital",
                "current": "Activo",
                "target": "100%",
                "progress": "on_track",
            },
            {
                "kr": f"Completar {total_jobs}+ registros en el período",
                "current": f"{total_jobs}",
                "target": "Creciente",
                "progress": "on_track" if total_jobs >= 20 else "needs_attention",
            },
            {
                "kr": f"Mantener {total_hours} horas hombre registradas con precisión",
                "current": f"{total_hours} hrs",
                "target": "≥100 hrs",
                "progress": "on_track" if total_hours >= 100 else "needs_attention",
            },
        ]
    })

    return okrs


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard_page(request: Request):
    return templates.TemplateResponse("dashboard.html", {"request": request})


@app.get("/api/dashboard/summary")
def dashboard_summary(
    date_from: str = Query(None),
    date_to: str = Query(None),
    group: str = Query(None),
    compare_from: str = Query(None),
    compare_to: str = Query(None),
    db: Session = Depends(get_db),
):
    try:
        today = date.today()
        if not date_from or not date_to:
            if today.day >= 26:
                period_start = today.replace(day=26)
            else:
                period_start = (today.replace(day=1) - timedelta(days=1)).replace(day=26)
            period_end = (period_start + timedelta(days=32)).replace(day=25)
            if not date_from:
                date_from = period_start.isoformat()
            if not date_to:
                date_to = period_end.isoformat()

        period_label = f"{parse_date(date_from).strftime('%d/%m/%Y')} – {parse_date(date_to).strftime('%d/%m/%Y')}"

        current = _compute_period_stats(db, date_from, date_to, group)
        current["label"] = period_label
        current["date_from"] = date_from
        current["date_to"] = date_to

        result = {
            "current": current,
            "comparison": None,
            "okrs": _compute_okrs(current),
        }

        if compare_from and compare_to:
            comp = _compute_period_stats(db, compare_from, compare_to, group)
            comp["label"] = f"{parse_date(compare_from).strftime('%d/%m/%Y')} – {parse_date(compare_to).strftime('%d/%m/%Y')}"
            comp["date_from"] = compare_from
            comp["date_to"] = compare_to
            result["comparison"] = comp

        return result
    except Exception as exc:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Error generando dashboard: {exc}")


@app.get("/api/dashboard/kpi-detail")
def kpi_detail(
    kpi: str = Query(...),
    date_from: str = Query(None),
    date_to: str = Query(None),
    group: str = Query(None),
    db: Session = Depends(get_db),
):
    try:
        q_base = _build_filtered_query(db, date_from, date_to, group)
        if kpi in ("pm_pct", "cm_pct"):
            action_filter = "Preventivo" if kpi == "pm_pct" else "Correctivo"
            rows = q_base.filter(JobEntry.action.ilike(f"%{action_filter}%")).order_by(JobEntry.id.desc()).limit(200).all()
            return {
                "title": f"Trabajos con acción «{action_filter}»",
                "headers": ["#", "Fecha", "Grupo", "Macroproceso", "Tipo", "Acción", "Equipo", "Duración"],
                "rows": [["N/A", str(e.report.date), e.report.group_name, e.macroprocess, e.work_type, e.action, e.equipment, e.duration] for e in rows],
            }
        if kpi == "worker_count":
            rows = (
                db.query(Worker.id, Worker.name, Worker.group_name)
                .join(Report)
            )
            if date_from:
                rows = rows.filter(Report.date >= parse_date(date_from))
            if date_to:
                rows = rows.filter(Report.date <= parse_date(date_to))
            if group:
                rows = rows.filter(Report.group_name == group)
            rows = rows.distinct().order_by(Worker.name).all()
            return {
                "title": "Trabajadores que reportaron",
                "headers": ["ID", "Nombre", "Grupo"],
                "rows": [[w.id, w.name, w.group_name] for w in rows],
            }
        if kpi == "collaborator_count":
            rows = q_base.filter(JobEntry.collaborators.isnot(None)).order_by(JobEntry.id.desc()).limit(300).all()
            names = set()
            data = []
            for e in rows:
                if e.collaborators:
                    for n in e.collaborators.split(","):
                        n = n.strip()
                        if n and n not in names:
                            names.add(n)
                            data.append([len(data) + 1, n])
            return {
                "title": "Colaboradores únicos en labor",
                "headers": ["#", "Nombre"],
                "rows": data,
            }
        if kpi == "total_hours":
            rows = q_base.filter(JobEntry.duration.isnot(None), JobEntry.duration != "").order_by(JobEntry.id.desc()).limit(300).all()
            return {
                "title": "Detalle de horas hombre",
                "headers": ["#", "Fecha", "Grupo", "Macroproceso", "Equipo", "Duración"],
                "rows": [[i + 1, str(e.report.date), e.report.group_name, e.macroprocess, e.equipment, e.duration] for i, e in enumerate(rows)],
            }
        # Default: list job entries (total_jobs, avg_duration_min, etc.)
        rows = q_base.order_by(JobEntry.id.desc()).limit(300).all()
        return {
            "title": "Todos los trabajos del período",
            "headers": ["#", "Fecha", "Grupo", "Macroproceso", "Tipo", "Acción", "Equipo", "Duración"],
            "rows": [[i + 1, str(e.report.date), e.report.group_name, e.macroprocess, e.work_type, e.action, e.equipment, e.duration] for i, e in enumerate(rows)],
        }
    except Exception as exc:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Error obteniendo detalle: {exc}")


# ─── Admin routes ─────────────────────────────────────────────────────────────

ADMIN_MODELS = {
    "turnos": Turno,
    "niveles": Nivel,
    "colaboradores": Collaborator,
    "macroprocesos": Macroprocess,
    "tipos_trabajo": WorkType,
    "acciones": Action,
    "equipos": EquipmentCategory,
    "subequipos": EquipmentSubitem,
}


ADMIN_PASSWORD = "Mantt.1"


@app.get("/admin", response_class=HTMLResponse)
def admin_page(request: Request):
    return templates.TemplateResponse("admin.html", {"request": request})


@app.post("/api/admin/verify")
def admin_verify(password: str = Body(..., embed=True)):
    if password == ADMIN_PASSWORD:
        return {"ok": True}
    raise HTTPException(401, "Contraseña incorrecta")


@app.get("/api/admin/{entity}")
def admin_list(entity: str, db: Session = Depends(get_db)):
    if entity == "equipos":
        cats = db.query(EquipmentCategory).order_by(EquipmentCategory.position).all()
        return [{"id": c.id, "name": c.name, "action_group": c.action_group, "subitems": [{"id": s.id, "name": s.name, "meters": s.meters} for s in c.subitems]} for c in cats]
    if entity == "subequipos":
        return db.query(EquipmentSubitem).order_by(EquipmentSubitem.position).all()
    model = ADMIN_MODELS.get(entity)
    if not model:
        raise HTTPException(400, "Entidad no valida")
    return db.query(model).order_by(model.position).all()


@app.post("/api/admin/{entity}")
def admin_create(entity: str, data: dict, db: Session = Depends(get_db)):
    if entity == "turnos":
        obj = Turno(name=data["name"], position=db.query(Turno).count())
    elif entity == "niveles":
        obj = Nivel(name=data["name"], position=db.query(Nivel).count())
    elif entity == "colaboradores":
        obj = Collaborator(group_name=data["group_name"], name=data["name"], position=db.query(Collaborator).count())
    elif entity == "macroprocesos":
        obj = Macroprocess(group_key=data["group_key"], name=data["name"], position=db.query(Macroprocess).count())
    elif entity == "tipos_trabajo":
        obj = WorkType(type_key=data["type_key"], name=data["name"], position=db.query(WorkType).count(), default_action=data.get("default_action", ""))
    elif entity == "acciones":
        obj = Action(group_key=data["group_key"], name=data["name"], position=db.query(Action).count())
    elif entity == "equipos":
        obj = EquipmentCategory(name=data["name"], action_group=data.get("action_group", ""), position=db.query(EquipmentCategory).count())
        db.add(obj)
        db.flush()
        if "subitems" in data:
            for i, s in enumerate(data["subitems"]):
                db.add(EquipmentSubitem(category_id=obj.id, name=s["name"], meters=s.get("meters", "fin"), position=i))
    elif entity == "subequipos":
        obj = EquipmentSubitem(category_id=data["category_id"], name=data["name"], meters=data.get("meters", "fin"), position=db.query(EquipmentSubitem).count())
    else:
        raise HTTPException(400, "Entidad no valida")
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@app.put("/api/admin/{entity}/{item_id}")
def admin_update(entity: str, item_id: int, data: dict, db: Session = Depends(get_db)):
    if entity == "equipos":
        obj = db.query(EquipmentCategory).filter(EquipmentCategory.id == item_id).first()
        if not obj: raise HTTPException(404)
        if "name" in data: obj.name = data["name"]
        if "action_group" in data: obj.action_group = data["action_group"]
        if "subitems" in data:
            db.query(EquipmentSubitem).filter(EquipmentSubitem.category_id == obj.id).delete()
            for i, s in enumerate(data["subitems"]):
                db.add(EquipmentSubitem(category_id=obj.id, name=s["name"], meters=s.get("meters", "fin"), position=i))
    elif entity == "subequipos":
        obj = db.query(EquipmentSubitem).filter(EquipmentSubitem.id == item_id).first()
        if not obj: raise HTTPException(404)
        if "name" in data: obj.name = data["name"]
        if "meters" in data: obj.meters = data["meters"]
    else:
        model = ADMIN_MODELS.get(entity)
        if not model: raise HTTPException(400)
        obj = db.query(model).filter(model.id == item_id).first()
        if not obj: raise HTTPException(404)
        for key, val in data.items():
            if hasattr(obj, key) and key not in ("id", "position"):
                setattr(obj, key, val)
    db.commit()
    db.refresh(obj)
    return obj


@app.put("/api/admin/{entity}/reorder/{item_id}")
def admin_reorder(entity: str, item_id: int, direction: str = "up", db: Session = Depends(get_db)):
    """Move an item up or down in position ordering."""
    import math

    if entity == "equipos":
        model_cls = EquipmentCategory
    elif entity == "subequipos":
        model_cls = EquipmentSubitem
    else:
        model_cls = ADMIN_MODELS.get(entity)
    if not model_cls:
        raise HTTPException(400, "Entidad no valida")

    item = db.query(model_cls).filter(model_cls.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item no encontrado")

    all_items = db.query(model_cls).order_by(model_cls.position, model_cls.id).all()

    idx = None
    for i, obj in enumerate(all_items):
        if obj.id == item_id:
            idx = i
            break
    if idx is None:
        raise HTTPException(404)

    swap_idx = idx - 1 if direction == "up" else idx + 1
    if swap_idx < 0 or swap_idx >= len(all_items):
        raise HTTPException(400, "Ya está al extremo")

    # Swap positions
    item.position, all_items[swap_idx].position = all_items[swap_idx].position, item.position
    db.commit()
    return {"ok": True}


@app.delete("/api/admin/{entity}/{item_id}")
def admin_delete(entity: str, item_id: int, db: Session = Depends(get_db)):
    if entity == "equipos":
        obj = db.query(EquipmentCategory).filter(EquipmentCategory.id == item_id).first()
    elif entity == "subequipos":
        obj = db.query(EquipmentSubitem).filter(EquipmentSubitem.id == item_id).first()
    else:
        model = ADMIN_MODELS.get(entity)
        if not model: raise HTTPException(400)
        obj = db.query(model).filter(model.id == item_id).first()
    if not obj: raise HTTPException(404)
    db.delete(obj)
    db.commit()
    return {"ok": True}


# ─── RACS (Reportes de Actos y Condiciones Subestándar) ──────────────────────

RACS_GROUPS = {
    "Trackless": ["Wilson Apaza","Elvis Avalos","Yoel Castillo","Ronaldo Ccompara","Yordan Elguera","Miguel Espiritu","Jordi Gallegos","Jimy Huaman","Rivaldo Mamani","Ademer Moreto","Eddy Quispe","Felix Quispe","Jesús Quispe","Jesús Ramirez","Edwin Torres","Martin Sillocca"],
    "Convencional": ["Moises de la Cadena","Edgardo del Carpio","Erik Inca","Alan Llanquecha","Diego Machaca","Vlady mamani","Artemio Payehuanca","Fredy Taya","Edgar Zela","Aldair Vilca","Abraham Dionisio"],
    "Electrico": ["Ronaldo Ccompara","Jhonatan Chipane","Koos Coaquira","Wilder Espinal","Markc Huachaca","Jose Huayra","Santiago Huayra","Julio Perez","Wilber Quijahuaman","Rronar Surco"],
}

# ─── Seed racs_workers from RACS_GROUPS (only if table empty) ─────────────
try:
    _d = SessionLocal()
    from .models import RacsWorker
    if _d.query(RacsWorker).count() == 0:
        for _g, _names in RACS_GROUPS.items():
            for _w in _names:
                _d.add(RacsWorker(name=_w, group_name=_g, active=True))
        _d.commit()
        print(f"[OK] RacsWorker seeded: {_d.query(RacsWorker).count()} workers")
    _d.close()
except Exception as _e:
    print(f"[WARN] RacsWorker seed: {_e}")

# ─── Seed guardias ───────────────────────────────────────────────────────
try:
    _d = SessionLocal()
    from .models import Guardia, WorkerGuardia
    if _d.query(Guardia).count() == 0:
        from datetime import date as _dd
        _d.add(Guardia(name="A", phase_start=_dd(2026, 7, 6)))
        _d.add(Guardia(name="B", phase_start=_dd(2026, 6, 26)))
        _d.add(Guardia(name="C", phase_start=_dd(2026, 6, 16)))
        _d.commit()
        print("[OK] Guardias seeded: A(06/07), B(26/06), C(16/06)")
    _d.close()
except Exception as _e:
    print(f"[WARN] Guardia seed: {_e}")


def _get_racs_period(dt=None):
    """Return (period_start, period_end) for the current RACS period.
    Period: Sunday 20:00 → next Sunday 18:00.
    """
    if dt is None:
        dt = datetime.now()
    days_back = (dt.weekday() + 1) % 7
    cand_start = (dt - timedelta(days=days_back)).replace(
        hour=20, minute=0, second=0, microsecond=0
    )
    if cand_start > dt:
        cand_start -= timedelta(days=7)
    cand_end = cand_start + timedelta(days=6, hours=22)
    if cand_end <= dt:
        cand_start = cand_end.replace(hour=20, minute=0)
        cand_end = cand_start + timedelta(days=6, hours=22)
    return cand_start, cand_end


def _is_guardia_on_site(guardia_name: str, check_date=None, db=None):
    """Return True if the guardia is currently working (on-site)."""
    if not guardia_name:
        return True
    if check_date is None:
        check_date = datetime.now().date()
    elif isinstance(check_date, datetime):
        check_date = check_date.date()
    if db is None:
        return True
    g = db.query(Guardia).filter(Guardia.name == guardia_name).first()
    if not g:
        return True
    elapsed = (check_date - g.phase_start).days
    if elapsed < 0:
        return True
    return (elapsed % 30) < 20


def _get_worker_guardias(db):
    """Return dict: (worker_name, group_name) -> guardia_name"""
    rows = (
        db.query(WorkerGuardia.worker_name, WorkerGuardia.group_name, Guardia.name)
        .join(Guardia, WorkerGuardia.guardia_id == Guardia.id)
        .all()
    )
    return {(r.worker_name, r.group_name): r.name for r in rows}


def _get_worker_cargos(db):
    """Return dict: (worker_name, group_name) -> cargo"""
    rows = db.query(RacsWorker.name, RacsWorker.group_name, RacsWorker.cargo).filter(RacsWorker.active == True).all()
    return {(r.name, r.group_name): r.cargo for r in rows}


def _get_all_workers(db):
    """Return list of dicts: {name, group, cargo} from DB."""
    q = db.query(RacsWorker).filter(RacsWorker.active == True)
    return [{"name": r.name, "group_name": r.group_name, "cargo": r.cargo} for r in q.all()]


@app.get("/racs", response_class=HTMLResponse)
def racs_form(request: Request):
    return templates.TemplateResponse("racs_form.html", {"request": request})


@app.get("/racs/dashboard", response_class=HTMLResponse)
def racs_dashboard(request: Request):
    return templates.TemplateResponse("racs_dashboard.html", {"request": request})


@app.get("/api/racs/period")
def racs_get_period():
    s, e = _get_racs_period()
    return {"period_start": s.isoformat(), "period_end": e.isoformat()}


@app.get("/api/racs/workers")
def racs_get_workers(db: Session = Depends(get_db)):
    """Return all workers grouped by team, with guardia and on-site status."""
    wg_map = _get_worker_guardias(db)
    cargos = _get_worker_cargos(db)
    all_workers = _get_all_workers(db)
    teams = {}
    for w in all_workers:
        teams.setdefault(w["group_name"], []).append(w)
    result = []
    for group, members in teams.items():
        workers = []
        for w in members:
            guardia_name = wg_map.get((w["name"], group), "")
            on_site = _is_guardia_on_site(guardia_name, db=db)
            workers.append({"name": w["name"], "guardia": guardia_name, "on_site": on_site, "cargo": w["cargo"] or ""})
        result.append({"group": group, "workers": workers})
    return result


@app.post("/api/racs", response_model=RacsReportResponse)
def racs_create(data: RacsReportCreate, db: Session = Depends(get_db)):
    ps, pe = _get_racs_period()
    from datetime import timezone
    import datetime as dt_mod
    fecha = None
    if data.fecha_reporte:
        try:
            fecha = dt_mod.datetime.strptime(data.fecha_reporte, "%Y-%m-%d").date()
        except:
            fecha = dt_mod.datetime.utcnow().date()
    r = RacsReport(
        worker_name=data.worker_name,
        group_name=data.group_name,
        categoria=data.categoria,
        tipo=data.tipo,
        turno=data.turno,
        descripcion=data.descripcion,
        ubicacion=data.ubicacion,
        referencia=data.referencia,
        nivel=data.nivel,
        fecha_reporte=fecha,
        riesgo=data.riesgo,
        accion_correctiva=data.accion_correctiva,
        tipo_descripcion=data.tipo_descripcion,
        period_start=ps,
        period_end=pe,
        created_at=dt_mod.datetime.now(timezone.utc),
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return r


@app.post("/api/racs/upload-photo/{racs_id}")
def racs_upload_photo(racs_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    r = db.query(RacsReport).filter(RacsReport.id == racs_id).first()
    if not r:
        raise HTTPException(404, "RACS no encontrado")
    ext = os.path.splitext(file.filename or "foto.jpg")[1] or ".jpg"
    fname = f"racs_{racs_id}_{uuid.uuid4().hex[:8]}{ext}"
    rel_dir = "racs_photos"
    dest_dir = os.path.join(STATIC_DIR, rel_dir)
    os.makedirs(dest_dir, exist_ok=True)
    dest = os.path.join(dest_dir, fname)
    with open(dest, "wb") as f:
        f.write(file.file.read())
    r.foto = f"/static/{rel_dir}/{fname}"
    db.commit()
    return {"ok": True, "path": r.foto}


@app.get("/api/racs/dashboard-data")
def racs_dashboard_data(db: Session = Depends(get_db)):
    ps, pe = _get_racs_period()
    wg_map = _get_worker_guardias(db)

    counts = (
        db.query(
            RacsReport.worker_name,
            RacsReport.group_name,
            _sf.count(RacsReport.id).label("cnt"),
        )
        .filter(RacsReport.period_start == ps, RacsReport.period_end == pe)
        .group_by(RacsReport.worker_name, RacsReport.group_name)
        .all()
    )
    count_map = {(r.worker_name, r.group_name): r.cnt for r in counts}

    all_workers = _get_all_workers(db)
    teams_dict = {}
    for w in all_workers:
        teams_dict.setdefault(w["group_name"], []).append(w["name"])
    teams = []
    total_workers = 0
    on_site_count = 0
    complete = 0
    partial = 0
    missing = 0
    for group, workers in teams_dict.items():
        members = []
        for w in workers:
            total_workers += 1
            guardia_name = wg_map.get((w, group), "")
            on_site = _is_guardia_on_site(guardia_name, db=db)
            c = count_map.get((w, group), 0)
            members.append({"name": w, "count": c, "guardia": guardia_name, "on_site": on_site})
            if on_site:
                on_site_count += 1
                if c >= 2: complete += 1
                elif c == 1: partial += 1
                else: missing += 1
        teams.append({"team": group, "workers": members})

    return {
        "period_start": ps.isoformat(),
        "period_end": pe.isoformat(),
        "teams": teams,
        "summary": {
            "total": total_workers,
            "on_site": on_site_count,
            "on_rest": total_workers - on_site_count,
            "complete": complete,
            "partial": partial,
            "missing": missing,
        },
    }


@app.get("/api/racs/dashboard-kpi")
def racs_dashboard_kpi(db: Session = Depends(get_db)):
    ps, pe = _get_racs_period()
    wg_map = _get_worker_guardias(db)

    all_reports = db.query(RacsReport).filter(
        RacsReport.period_start == ps, RacsReport.period_end == pe
    ).all()

    total = len(all_reports)

    # Counts by dimension
    cat_count = {"Seguridad y Salud Ocupacional": 0, "Medio Ambiente": 0}
    tipo_count = {"Acto Subestándar": 0, "Condición Subestándar": 0}
    riesgo_count = {"Alto": 0, "Medio": 0, "Bajo": 0}
    turno_count = {"DÍA": 0, "NOCHE": 0}
    nivel_count = {}
    guardia_count = {}
    daily_count = {}
    worker_racs = {}  # worker -> {count, category_counts, tipo_counts, riesgo_counts}

    for r in all_reports:
        cat = r.categoria or "Sin categoría"
        if cat in cat_count: cat_count[cat] += 1
        tp = r.tipo or "Sin tipo"
        if tp in tipo_count: tipo_count[tp] += 1
        rg = r.riesgo or "Sin riesgo"
        if rg in riesgo_count: riesgo_count[rg] += 1
        tn = r.turno or "Sin turno"
        if tn in turno_count: turno_count[tn] += 1

        nv = r.nivel or "Sin nivel"
        nivel_count[nv] = nivel_count.get(nv, 0) + 1

        g = wg_map.get((r.worker_name, r.group_name), "Sin guardia")
        guardia_count[g] = guardia_count.get(g, 0) + 1

        day_key = r.created_at.strftime("%Y-%m-%d")
        daily_count[day_key] = daily_count.get(day_key, 0) + 1

        wk = (r.worker_name, r.group_name)
        if wk not in worker_racs:
            worker_racs[wk] = {"count": 0, "cat": {}, "tipo": {}, "riesgo": {}, "guardia": g}
        worker_racs[wk]["count"] += 1
        worker_racs[wk]["cat"][cat] = worker_racs[wk]["cat"].get(cat, 0) + 1
        worker_racs[wk]["tipo"][tp] = worker_racs[wk]["tipo"].get(tp, 0) + 1
        worker_racs[wk]["riesgo"][rg] = worker_racs[wk]["riesgo"].get(rg, 0) + 1

    # Worker compliance
    all_workers = _get_all_workers(db)
    active_workers = 0
    for w in all_workers:
        g = wg_map.get((w["name"], w["group_name"]), "")
        if _is_guardia_on_site(g, db=db):
            active_workers += 1

    # Sort daily trend
    daily_trend = [{"date": k, "count": v} for k, v in sorted(daily_count.items())]

    # Worker detail for compliance
    worker_detail = []
    for (name, group), data in worker_racs.items():
        g = data["guardia"]
        on_site = _is_guardia_on_site(g, db=db)
        worker_detail.append({
            "name": name,
            "group": group,
            "count": data["count"],
            "guardia": g,
            "on_site": on_site,
            "riesgos": data["riesgo"],
        })

    # Sort nivel_count
    nivel_sorted = [{"nivel": k, "count": v} for k, v in sorted(nivel_count.items(), key=lambda x: -x[1])]

    return {
        "total_racs": total,
        "active_workers": active_workers,
        "period_start": ps.isoformat(),
        "period_end": pe.isoformat(),
        "by_categoria": [{"categoria": k, "count": v} for k, v in cat_count.items()],
        "by_tipo": [{"tipo": k, "count": v} for k, v in tipo_count.items()],
        "by_riesgo": [{"riesgo": k, "count": v} for k, v in riesgo_count.items()],
        "by_turno": [{"turno": k, "count": v} for k, v in turno_count.items()],
        "by_nivel": nivel_sorted,
        "by_guardia": [{"guardia": k, "count": v} for k, v in sorted(guardia_count.items())],
        "daily_trend": daily_trend,
        "worker_detail": worker_detail,
    }


@app.get("/api/racs/list")
def racs_list(db: Session = Depends(get_db)):
    ps, pe = _get_racs_period()
    rows = (
        db.query(RacsReport)
        .filter(RacsReport.period_start == ps, RacsReport.period_end == pe)
        .order_by(RacsReport.created_at.desc())
        .all()
    )
    return [
        {
            "id": r.id,
            "worker_name": r.worker_name,
            "group_name": r.group_name,
            "tipo": r.tipo,
            "categoria": r.categoria,
            "turno": r.turno,
            "descripcion": r.descripcion,
            "ubicacion": r.ubicacion,
            "referencia": r.referencia,
            "nivel": r.nivel,
            "fecha_reporte": (r.fecha_reporte.isoformat() if r.fecha_reporte else None),
            "riesgo": r.riesgo,
            "accion_correctiva": r.accion_correctiva,
            "tipo_descripcion": r.tipo_descripcion,
            "foto": r.foto,
            "created_at": (r.created_at.isoformat() if r.created_at.tzinfo else r.created_at.isoformat() + "Z"),
        }
        for r in rows
    ]


@app.get("/api/racs/worker-status")
def racs_worker_status(worker_name: str, db: Session = Depends(get_db)):
    """Return RACS count for a specific worker in current period."""
    ps, pe = _get_racs_period()
    cnt = (
        db.query(_sf.count(RacsReport.id))
        .filter(
            RacsReport.worker_name == worker_name,
            RacsReport.period_start == ps,
            RacsReport.period_end == pe,
        )
        .scalar()
    ) or 0
    return {"worker_name": worker_name, "count": cnt, "required": 2}


# ─── RACS Excel download ─────────────────────────────────────────────────────


TEMPLATE_RACS = os.path.join(os.path.dirname(__file__), "templates", "racs_template.xlsx")


@app.get("/api/racs/{racs_id}/excel")
def racs_download_excel(racs_id: int, db: Session = Depends(get_db)):
    r = db.query(RacsReport).filter(RacsReport.id == racs_id).first()
    if not r:
        raise HTTPException(404, "RACS no encontrado")
    if not os.path.exists(TEMPLATE_RACS):
        raise HTTPException(500, "Plantilla RACS no encontrada")

    wb = openpyxl.load_workbook(TEMPLATE_RACS)
    ws = wb["RACS"]

    def _set_cell(coord, value):
        for mr in list(ws.merged_cells.ranges):
            if coord in mr:
                ws.unmerge_cells(str(mr))
                break
        ws[coord] = value

    # Datos Generales
    _set_cell("C6", f"Reportado por: {r.worker_name}")
    _set_cell("P6", f"Cargo: Operador")
    fecha_str = r.fecha_reporte.strftime('%d/%m/%Y') if r.fecha_reporte else r.created_at.strftime('%d/%m/%Y')
    _set_cell("C7", f"Fecha: {fecha_str}")
    _set_cell("J7", f"Turno: {r.turno or 'DÍA'}")
    _set_cell("O7", f"Hora: {r.created_at.strftime('%H:%M')}")
    g_name = ""
    wg = _get_worker_guardias(db)
    guardia_name = wg.get((r.worker_name, r.group_name), "")
    if guardia_name:
        g_obj = db.query(Guardia).filter(Guardia.name == guardia_name).first()
        if g_obj:
            d = (datetime.now().date() - g_obj.phase_start).days
            cycle = d % 30 if d >= 0 else 0
            g_name = f"{guardia_name} (día {cycle+1}/30)" if cycle < 20 else f"{guardia_name} (descanso)"
    _set_cell("T7", f"Guardia: {g_name or guardia_name}")
    _set_cell("C8", "Empresa: U.M. Soledad")
    _set_cell("K8", f"Nivel: {r.nivel or ''}")
    _set_cell("O8", f"Labor / Lugar: {r.ubicacion or ''}")

    # Tipo checkboxes
    if r.tipo == "Acto Subestándar":
        _set_cell("N10", "X  Acto Subestándar")
        ws["N10"].font = openpyxl.styles.Font(bold=True, color="FF0000", size=10)
    else:
        _set_cell("S10", "X  Condición Subestándar")
        ws["S10"].font = openpyxl.styles.Font(bold=True, color="FF0000", size=10)

    # Descripción del Reporte (write below the title)
    _set_cell("C14", r.descripcion or "")

    # Acción Correctiva
    _set_cell("C21", r.accion_correctiva or "")

    # Mark the tipo_descripcion in the checklist (rows 25-38)
    if r.tipo_descripcion:
        td = r.tipo_descripcion
        # td format: "15. Falta accesorios / insumos / herramientas"
        parts = td.split(". ", 1)
        num_part = parts[0] if parts else ""
        desc_part = parts[1] if len(parts) > 1 else ""
        # Search in checklist rows
        for row in range(25, 39):
            for col_letter in ['D', 'K', 'R']:
                cell = ws[f'{col_letter}{row}']
                if cell.value and desc_part and desc_part.strip().lower() in str(cell.value).strip().lower():
                    col_idx = openpyxl.utils.column_index_from_string(col_letter) - 1
                    prev_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws[f'{col_letter}{row}'].font = openpyxl.styles.Font(bold=True, color="FF0000", size=9)
                    ws[f'{prev_letter}{row}'].font = openpyxl.styles.Font(bold=True, color="FF0000", size=9)

    # Risk level
    riesgo_mapping = {"Alto": "C40", "Medio": "E40", "Bajo": "G40"}
    if r.riesgo and r.riesgo in riesgo_mapping:
        coord = riesgo_mapping[r.riesgo]
        _set_cell(coord, f"X  {r.riesgo}")
        ws[coord].font = openpyxl.styles.Font(bold=True, color="FF0000", size=9)

    # Describe el tipo de reporte
    _set_cell("J40", f"Describe el tipo de reporte: {r.tipo_descripcion or ''}")

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"RACS_{r.id}_{r.worker_name.replace(' ','_')}_{r.created_at.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# ─── Guardia admin endpoints ──────────────────────────────────────────────────


@app.get("/api/guardias")
def guardia_list(db: Session = Depends(get_db)):
    guardias = db.query(Guardia).order_by(Guardia.id).all()
    wg = _get_worker_guardias(db)
    cargos = _get_worker_cargos(db)
    all_workers = _get_all_workers(db)
    teams_dict = {}
    for w in all_workers:
        teams_dict.setdefault(w["group_name"], []).append(w["name"])
    workers = [
        {"group": g, "workers": [{"name": w, "guardia": wg.get((w, g), ""), "cargo": cargos.get((w, g), "")} for w in names]}
        for g, names in teams_dict.items()
    ]
    return {
        "guardias": [{"id": g.id, "name": g.name, "phase_start": g.phase_start.isoformat()} for g in guardias],
        "workers": workers,
    }


@app.post("/api/guardias/update")
def guardia_update(data: dict = Body(...), db: Session = Depends(get_db)):
    """Update guardia phase_start or worker guardia assignment. Protected by DNI."""
    from datetime import date as _dd, datetime as _dt

    action = data.get("action")
    if action == "update_phase":
        g = db.query(Guardia).filter(Guardia.name == data["guardia"]).first()
        if g:
            g.phase_start = _dd.fromisoformat(data["phase_start"])
            db.commit()
            return {"ok": True}
    elif action == "assign_worker":
        existing = db.query(WorkerGuardia).filter(
            WorkerGuardia.worker_name == data["worker_name"],
            WorkerGuardia.group_name == data["group_name"],
        ).first()
        if data.get("guardia"):
            guardia = db.query(Guardia).filter(Guardia.name == data["guardia"]).first()
            if not guardia:
                raise HTTPException(400, "Guardia no encontrada")
            if existing:
                existing.guardia_id = guardia.id
            else:
                db.add(WorkerGuardia(worker_name=data["worker_name"], group_name=data["group_name"], guardia_id=guardia.id))
        elif existing:
            db.delete(existing)
        if data.get("cargo") is not None and existing:
            existing.cargo = data["cargo"]
        if data.get("cargo") is not None:
            rw = db.query(RacsWorker).filter(RacsWorker.name == data["worker_name"], RacsWorker.group_name == data["group_name"], RacsWorker.active == True).first()
            if rw:
                rw.cargo = data["cargo"]
        db.commit()
        return {"ok": True}
    raise HTTPException(400, "Acción no válida")


# ─── RACS Worker CRUD endpoints ────────────────────────────────────────────


@app.get("/api/racs/workers/list")
def racs_workers_list(db: Session = Depends(get_db)):
    wg = _get_worker_guardias(db)
    all_workers = _get_all_workers(db)
    teams = {}
    for w in all_workers:
        teams.setdefault(w["group_name"], []).append(w)
    result = []
    for group, members in teams.items():
        workers = []
        for w in members:
            guardia_name = wg.get((w["name"], group), "")
            workers.append({"name": w["name"], "cargo": w["cargo"] or "", "guardia": guardia_name, "group": group})
        result.append({"group": group, "workers": workers})
    return result


@app.post("/api/racs/workers/create")
def racs_worker_create(data: RacsWorkerCreate, db: Session = Depends(get_db)):
    existing = db.query(RacsWorker).filter(RacsWorker.name == data.name, RacsWorker.group_name == data.group_name).first()
    if existing:
        if not existing.active:
            existing.active = True
            existing.cargo = data.cargo or existing.cargo
            db.commit()
            return {"ok": True, "id": existing.id}
        raise HTTPException(400, "El trabajador ya existe en este equipo")
    obj = RacsWorker(name=data.name, group_name=data.group_name, cargo=data.cargo, active=True)
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return {"ok": True, "id": obj.id}


@app.put("/api/racs/workers/{worker_id}")
def racs_worker_update(worker_id: int, data: RacsWorkerUpdate, db: Session = Depends(get_db)):
    obj = db.query(RacsWorker).filter(RacsWorker.id == worker_id).first()
    if not obj:
        raise HTTPException(404)
    old_name = obj.name
    old_group = obj.group_name
    if data.name is not None: obj.name = data.name
    if data.group_name is not None: obj.group_name = data.group_name
    if data.cargo is not None: obj.cargo = data.cargo
    if data.active is not None: obj.active = data.active
    # Sync WorkerGuardia if name or group changed
    if (data.name is not None or data.group_name is not None) and (old_name != obj.name or old_group != obj.group_name):
        wg = db.query(WorkerGuardia).filter(
            WorkerGuardia.worker_name == old_name,
            WorkerGuardia.group_name == old_group
        ).first()
        if wg:
            wg.worker_name = obj.name
            wg.group_name = obj.group_name
    if data.guardia is not None:
        existing_wg = db.query(WorkerGuardia).filter(
            WorkerGuardia.worker_name == obj.name,
            WorkerGuardia.group_name == obj.group_name,
        ).first()
        if data.guardia:
            guardia = db.query(Guardia).filter(Guardia.name == data.guardia).first()
            if not guardia:
                raise HTTPException(400, "Guardia no encontrada")
            if existing_wg:
                existing_wg.guardia_id = guardia.id
            else:
                db.add(WorkerGuardia(worker_name=obj.name, group_name=obj.group_name, guardia_id=guardia.id))
        elif existing_wg:
            db.delete(existing_wg)
    db.commit()
    db.refresh(obj)
    return {"ok": True, "id": obj.id, "name": obj.name, "group_name": obj.group_name, "cargo": obj.cargo}


@app.post("/api/racs/workers/change-group")
def racs_worker_change_group(data: dict = Body(...), db: Session = Depends(get_db)):
    name = data.get("name")
    old_group = data.get("old_group")
    new_group = data.get("new_group")
    if not name or not old_group or not new_group:
        raise HTTPException(400, "Faltan datos")
    obj = db.query(RacsWorker).filter(RacsWorker.name == name, RacsWorker.group_name == old_group, RacsWorker.active == True).first()
    if not obj:
        raise HTTPException(404, "Trabajador no encontrado")
    obj.group_name = new_group
    # Update WorkerGuardia if exists
    wg = db.query(WorkerGuardia).filter(WorkerGuardia.worker_name == name, WorkerGuardia.group_name == old_group).first()
    if wg:
        wg.group_name = new_group
    db.commit()
    return {"ok": True}


@app.get("/api/racs/groups")
def racs_get_groups(db: Session = Depends(get_db)):
    """Return distinct group names from active RACS workers."""
    q = db.query(RacsWorker.group_name).filter(RacsWorker.active == True).distinct().order_by(RacsWorker.group_name).all()
    return [g[0] for g in q]


@app.put("/api/racs/workers/{worker_id}")
def racs_worker_update(worker_id: int, data: RacsWorkerUpdate, db: Session = Depends(get_db)):
    obj = db.query(RacsWorker).filter(RacsWorker.id == worker_id).first()
    if not obj:
        raise HTTPException(404, "Trabajador no encontrado")
    old_name = obj.name
    old_group = obj.group_name
    if data.name is not None:
        obj.name = data.name
    if data.group_name is not None:
        obj.group_name = data.group_name
    if data.cargo is not None:
        obj.cargo = data.cargo
    # Sync WorkerGuardia if name or group changed
    if data.name is not None or data.group_name is not None:
        wg = db.query(WorkerGuardia).filter(
            WorkerGuardia.worker_name == old_name,
            WorkerGuardia.group_name == old_group
        ).first()
        if wg:
            wg.worker_name = obj.name
            wg.group_name = obj.group_name
    db.commit()
    db.refresh(obj)
    return {"ok": True, "id": obj.id, "name": obj.name, "group_name": obj.group_name, "cargo": obj.cargo}


@app.delete("/api/racs/workers/delete-by-name")
def racs_worker_delete_by_name(name: str = Query(...), group: str = Query(...), db: Session = Depends(get_db)):
    obj = db.query(RacsWorker).filter(RacsWorker.name == name, RacsWorker.group_name == group, RacsWorker.active == True).first()
    if not obj:
        raise HTTPException(404, "Trabajador no encontrado")
    obj.active = False
    db.commit()
    return {"ok": True}


@app.delete("/api/racs/workers/{worker_id}")
def racs_worker_delete(worker_id: int, db: Session = Depends(get_db)):
    obj = db.query(RacsWorker).filter(RacsWorker.id == worker_id).first()
    if not obj:
        raise HTTPException(404)
    obj.active = False
    db.commit()
    return {"ok": True}


# ─── RACS database Excel export ──────────────────────────────────────────


RACS_DB_TEMPLATE = os.path.join(os.path.dirname(__file__), "templates", "racs_db_template.xlsx")


@app.get("/api/racs/database-excel")
def racs_database_excel(db: Session = Depends(get_db)):
    from openpyxl.styles import Font as XlFont, Alignment as XlAlign, PatternFill as XlFill, Border as XlBorder, Side as XlSide
    from openpyxl.utils import get_column_letter

    reports = db.query(RacsReport).order_by(RacsReport.created_at.asc()).all()
    wg_map = _get_worker_guardias(db)
    cargo_map = _get_worker_cargos(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RACS GENERAL PROSOL SA"

    # ── Styles ──
    hdr_fill = XlFill("solid", fgColor="000000")
    hdr_font = XlFont(bold=True, color="FFFFFF", size=10, name="Calibri")
    hdr_align = XlAlign(horizontal="center", vertical="center", wrap_text=True)
    title_font = XlFont(bold=True, size=14, name="Calibri")
    sub_font = XlFont(size=8, name="Calibri")
    cell_font = XlFont(size=9, name="Calibri")
    cell_align = XlAlign(vertical="top", wrap_text=True)
    thin_side = XlSide(style="thin", color="000000")
    thin_border = XlBorder(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ── Title ──
    ws.merge_cells("A1:B2")
    ws["A1"].font = title_font

    ws.merge_cells("C1:R2")
    ws["C1"].value = "BASE DE DATOS DE REPORTE DE ACTO Y CONDICIONES SUBESTANDARES - RACS CIA PROSOL ISPACAS S.A."
    ws["C1"].font = title_font
    ws["C1"].alignment = XlAlign(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("S1:S2")
    ws["S1"].font = sub_font

    ws.merge_cells("T2:W2")
    ws["T2"].value = ","
    ws["T2"].font = sub_font
    ws.merge_cells("T3:W3")
    ws["T3"].font = sub_font
    ws.merge_cells("T4:W4")
    ws["T4"].font = sub_font

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # ── Headers row 5 ──
    headers = [
        ("A", "CODIGO", 8),
        ("B", "SEMANA", 10),
        ("C", "MES", 12),
        ("D", "FECHA DE REPORTE", 14),
        ("E", "AREA", 10),
        ("F", "TIPO", 12),
        ("G", "GUARDIA", 10),
        ("H", "TURNO", 10),
        ("I", "NOMBRE DEL REPORTANTE", 28),
        ("J", "CARGO", 22),
        ("K", "NIVEL", 8),
        ("L", "LUGAR", 16),
        ("M", "Descripción de la Condición o Acto", 35),
        ("N", "Acción a implementar", 35),
        ("O", "Tipo de Causa", 25),
        ("P", "Controles Criticos", 18),
        ("Q", "Riesgo\n(A,M, B)", 12),
        ("R", "DESCRIPCION DE MEDIDA\n(EJECUTADO, EN PROCESO o PENDIENTE)", 22),
        ("S", "EMPRESA REPORTANTE", 18),
        ("T", "EMPRESA RESPONSABLE", 18),
        ("U", "PROCESO DE AVANCE", 18),
    ]

    for col_letter, name, width in headers:
        col_idx = openpyxl.utils.column_index_from_string(col_letter)
        cell = ws.cell(row=5, column=col_idx, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:U5"

    # ── Data ──
    meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
             "JULIO", "AGOSTO", "SETIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

    for idx, r in enumerate(reports, 1):
        row_num = 5 + idx
        dt_local = r.created_at
        fecha_reporte = r.fecha_reporte or dt_local
        mes = meses[dt_local.month - 1]
        semana = f"SEM {dt_local.isocalendar()[1]:02d}"
        guardia = wg_map.get((r.worker_name, r.group_name), "")
        cargo = cargo_map.get((r.worker_name, r.group_name), "")

        values = [
            idx, semana, mes, fecha_reporte.strftime("%d/%m/%Y"), "MINA",
            r.tipo, guardia, r.turno or "DÍA",
            r.worker_name, cargo, r.nivel or "",
            f"{r.ubicacion or ''} / Ref: {r.referencia or ''}",
            r.descripcion or "", r.accion_correctiva or "", r.tipo_descripcion or "",
            "", r.riesgo or "", "EJECUTADO",
            "Inversiones Prosol", "", "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="BD_Acto_Condicion_Subestandar_{datetime.now().strftime("%Y%m%d")}.xlsx"'},
    )


# ─── Auto-scheduler for daily reports ──────────────────────────────────────
#
# Uncomment the start_scheduler() call below and set reload=False in run.py
# to enable automatic daily report generation at SCHEDULE_HOUR:00.
#
# import threading
# import time as time_module
#
# SCHEDULE_HOUR = 6  # 6:00 AM
# SCHEDULE_MINUTE = 0
#
#
# def _get_db_session():
#     from .database import SessionLocal as SL
#     return SL()
#
#
# def _auto_generate():
#     from datetime import timedelta
#     yesterday = (datetime.now() - timedelta(days=1)).date()
#     fname = f"Reporte_Diario_{yesterday.isoformat()}.xlsx"
#     fpath = os.path.join(REPORTS_DIR, fname)
#     if os.path.exists(fpath):
#         return
#     db = _get_db_session()
#     try:
#         reports = db.query(Report).options(
#             joinedload(Report.worker), joinedload(Report.entries)
#         ).filter(Report.date == yesterday).all()
#         if not reports:
#             return
#         grouped = {}
#         for rp in reports:
#             group = rp.group_name
#             shift = rp.shift
#             grouped.setdefault(group, {}).setdefault(shift, []).append({
#                 "id": rp.id, "date": rp.date, "shift": shift, "group_name": group,
#                 "collaborators_trackless": rp.collaborators_trackless,
#                 "collaborators_convencional": rp.collaborators_convencional,
#                 "collaborators_electrico": rp.collaborators_electrico,
#                 "entries": [{"macroprocess": e.macroprocess, "work_type": e.work_type, "work_subtype": e.work_subtype, "action": e.action, "description": e.description, "level": e.level, "location": e.location, "start_time_int": e.start_time_int, "end_time_int": e.end_time_int, "duration": e.duration, "equipment": e.equipment, "horometer_motor": e.horometer_motor, "horometer_motor_jumbo": e.horometer_motor_jumbo, "horometer_motor_volquetes": e.horometer_motor_volquetes, "horometer_electric": e.horometer_electric, "horometer_percussion": e.horometer_percussion, "kilometer": e.kilometer} for e in rp.entries],
#             })
#         generate_daily_report(yesterday, grouped)
#     except Exception as exc:
#         print(f"[Scheduler] Error generating daily report: {exc}")
#     finally:
#         db.close()
#
#
# def _scheduler_loop():
#     while True:
#         now = datetime.now()
#         if now.hour == SCHEDULE_HOUR and now.minute == SCHEDULE_MINUTE:
#             _auto_generate()
#             time_module.sleep(61)
#         time_module.sleep(55)
#
#
# def start_scheduler():
#     t = threading.Thread(target=_scheduler_loop, daemon=True)
#     t.start()
#
#
# # To enable auto-scheduling, uncomment the next line and set reload=False in run.py:
# # start_scheduler()
