"""
Daily report generator — builds Excel workbook from the reference template,
preserving all images, merged cells, colors, and formatting.
"""

import os
import io
import copy as _copy
from datetime import date
from typing import List, Optional
from openpyxl import load_workbook as _load_wb
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage

from .excel_export import (
    clean_activity, standardize_time, calc_minutes,
    extract_equipment, combine_horometer,
)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_DIR = os.path.join(BASE_DIR, "app", "templates")

TEMPLATE_PATH = os.path.join(TEMPLATE_DIR, "Reporte_Diario_Mantt.xlsx")
FOOTER_IMAGE_PATH = os.path.join(TEMPLATE_DIR, "firma_footer.png")

DATA_FONT = Font(name="Calibri", size=12, color="000000")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="000000")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ROW_EVEN_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _copy_sheet(wb, template_ws, new_title, idx):
    """Copy worksheet (copy_worksheet does NOT copy images)."""
    new_ws = wb.copy_worksheet(template_ws)
    new_ws.title = new_title
    if idx != wb.index(new_ws):
        wb.move_sheet(new_ws, offset=idx - wb.index(new_ws))
    return new_ws


def _collect_images(ws):
    """Extract (image_bytes, anchor_obj, width, height) from a worksheet."""
    items = []
    for img in getattr(ws, '_images', []) or []:
        try:
            ref = img.ref
            if isinstance(ref, io.BytesIO):
                ref.seek(0)
                data = ref.read()
                items.append((data, img.anchor, img.width, img.height))
            else:
                data = ref.read() if hasattr(ref, 'read') else None
                if data:
                    items.append((data, img.anchor, img.width, img.height))
        except Exception:
            pass
    return items


def _restore_images(ws, items):
    """Add previously collected images to a worksheet."""
    for data_bytes, anchor_obj, width, height in items:
        try:
            img = XlImage(io.BytesIO(data_bytes))
            img.anchor = anchor_obj
            img.width = width
            img.height = height
            ws.add_image(img)
        except Exception:
            pass


def _ensure_data_rows(ws, needed: int):
    """Insert extra rows if more than 10 data rows are needed."""
    data_start = 15
    template_count = 10
    if needed <= template_count:
        return
    extra = needed - template_count
    ws.insert_rows(25, extra)
    for i in range(template_count, needed):
        row = data_start + i
        ws.row_dimensions[row].height = 50
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=17)


def _calc_row_height(d: dict) -> float:
    max_lines = 1
    names = [n for n in d.get("trabajadores", []) if n]
    if names:
        max_lines = max(max_lines, len(names))
    acts = d.get("actividades", "") or ""
    if acts:
        max_lines = max(max_lines, len(acts) // 40 + 1)
    for key in ("macro", "tipo_servicio", "accion"):
        val = d.get(key, "") or ""
        if len(val) > 35:
            max_lines = max(max_lines, len(val) // 35 + 1)
    return min(max(max_lines * 16, 50), 200)


def _populate_data(ws, data_rows: list, header_info: dict):
    ws["B6"].value = "FECHA:"
    ws["D6"].value = header_info["date"]
    ws["B7"].value = "TURNO"
    ws["D7"].value = header_info["shift"]
    ws["B10"].value = "AREA"
    ws["D10"].value = header_info["group_name"]

    num_rows = max(len(data_rows), 10)

    # Remove extra rows left from previous template copies
    current_max = ws.max_row
    if current_max > 25:
        ws.delete_rows(25, current_max - 24)

    from openpyxl.utils import range_boundaries
    for merge_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
        if min_row > 24:
            ws.unmerge_cells(str(merge_range))

    _ensure_data_rows(ws, num_rows)

    data_start = 15
    for i in range(num_rows):
        row = data_start + i
        ws.row_dimensions[row].height = 50

        for col in range(1, 24):
            c = ws.cell(row=row, column=col)
            from openpyxl.cell.cell import MergedCell
            if isinstance(c, MergedCell):
                continue
            c.value = None
            c.fill = PatternFill(fill_type=None)
            c.font = DATA_FONT
            c.alignment = LEFT_CENTER
            c.border = BORDER_THIN

        if i < len(data_rows):
            d = data_rows[i]
            ws.cell(row=row, column=1).value = i + 1
            ws.cell(row=row, column=1).alignment = CENTER

            names = [n for n in d.get("trabajadores", []) if n]
            if names:
                ws.cell(row=row, column=2).value = "\n".join(f"{j+1}. {n}" for j, n in enumerate(names))

            ws.cell(row=row, column=6).value = d.get("macro", "")
            ws.cell(row=row, column=7).value = d.get("tipo_servicio", "")
            ws.cell(row=row, column=8).value = d.get("accion", "")
            ws.cell(row=row, column=9).value = d.get("actividades", "")
            ws.cell(row=row, column=18).value = d.get("equipo", "")
            ws.cell(row=row, column=19).value = d.get("equipo_interv", "")

            minutos = d.get("minutos")
            if minutos is not None:
                ws.cell(row=row, column=21).value = minutos

            ws.cell(row=row, column=22).value = d.get("nivel", "")
            ws.cell(row=row, column=23).value = d.get("labor", "")

            ws.row_dimensions[row].height = _calc_row_height(d)

        for col in range(1, 24):
            c = ws.cell(row=row, column=col)
            from openpyxl.cell.cell import MergedCell
            if isinstance(c, MergedCell):
                continue
            if c.value is None:
                c.value = ""
            c.font = DATA_FONT
            c.alignment = LEFT_CENTER
            c.border = BORDER_THIN
            if i % 2 == 1:
                c.fill = ROW_EVEN_FILL

    for row_num in range(data_start, data_start + num_rows):
        c = ws.cell(row=row_num, column=1)
        c.number_format = '0'


def _add_footer_images(wb, sheet_num_rows: dict):
    if not os.path.exists(FOOTER_IMAGE_PATH):
        return
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        n = sheet_num_rows.get(sheet_name, 0)
        footer_row = max(28, 15 + max(n, 10) + 1)
        try:
            img = XlImage(FOOTER_IMAGE_PATH)
            img.anchor = f"A{footer_row}"
            img.width = 2351
            img.height = 47
            ws.add_image(img)
        except Exception:
            pass


def _build_workbook(target_date: date, reports_data: dict):
    """Build workbook in memory and return (workbook, output_filename)."""
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template no encontrado: {TEMPLATE_PATH}")

    wb = _load_wb(TEMPLATE_PATH)
    template_ws = wb["Report_Diario"]

    groups = ["Mantt. Eq. Trackless", "Mantt. Eq. Convencional", "Mantt. Eq. Electrico"]
    shifts = ["Día", "Noche"]
    shift_aliases = {"Día": ["Día", "Dia", "dia", "día"], "Noche": ["Noche", "noche"]}
    date_str = target_date.strftime("%d/%m/%Y")
    short_date = target_date.strftime("%d%m%Y")

    template_ws.title = "__template__"
    src_ws = wb["__template__"]

    # Collect template images BEFORE any modifications (copy_worksheet loses them)
    template_images = _collect_images(src_ws)

    sheet_count = 0
    sheet_names = []
    sheet_num_rows = {}

    for group in groups:
        for shift in shifts:
            group_data = reports_data.get(group, {})
            shift_key = shift
            for alias in shift_aliases.get(shift, [shift]):
                if alias in group_data:
                    shift_key = alias
                    break
            shift_data = group_data.get(shift_key, [])
            if not shift_data:
                continue

            all_entries = []
            for rp in shift_data:
                for entry in (rp.get("entries", []) or [None]):
                    entry = entry or {}
                    start_str = standardize_time(entry.get("start_time_int"))
                    end_str = standardize_time(entry.get("end_time_int"))
                    dur_raw = entry.get("duration")
                    try:
                        dur_min = float(dur_raw) if dur_raw is not None else None
                    except (ValueError, TypeError):
                        dur_min = None
                    minutos = calc_minutes(start_str, end_str, rp.get("shift"), dur_min)
                    equip_raw = entry.get("equipment", "") or ""
                    equip_interv, equipo_code = extract_equipment(equip_raw)
                    equipo = equipo_code or equip_interv
                    actividades = clean_activity(entry.get("description"))
                    macro = entry.get("macroprocess", "")
                    tipo_serv = entry.get("work_type", "")
                    if entry.get("work_subtype"):
                        tipo_serv = f"{tipo_serv} - {entry['work_subtype']}"
                    accion = entry.get("action", "")
                    nivel = entry.get("level", "")
                    labor = entry.get("location", "")

                    raw_colabs = entry.get("collaborators") or ""
                    colab_parts = [c.strip() for c in raw_colabs.split(",") if c.strip()]
                    trabajadores = []
                    for j in range(4):
                        trabajadores.append(colab_parts[j] if j < len(colab_parts) else None)

                    all_entries.append({
                        "trabajadores": trabajadores,
                        "actividades": actividades,
                        "macro": macro,
                        "tipo_servicio": tipo_serv,
                        "accion": accion,
                        "equipo_interv": equip_interv,
                        "equipo": equipo,
                        "minutos": minutos,
                        "nivel": nivel,
                        "labor": labor,
                    })

            if not all_entries:
                continue

            safe_group = group.replace("Mantt. Eq. ", "").replace(" ", "_")
            sheet_name = f"{safe_group}_{shift}_{short_date}"[:31]

            header_info = {"date": date_str, "shift": shift, "group_name": group}

            ws = _copy_sheet(wb, src_ws, sheet_name, sheet_count)
            _populate_data(ws, all_entries, header_info)
            sheet_names.append(sheet_name)
            sheet_num_rows[sheet_name] = len(all_entries)
            sheet_count += 1

    if sheet_count == 0:
        ws = _copy_sheet(wb, src_ws, f"Sin_Datos_{short_date}"[:31], 0)
        _populate_data(ws, [], {"date": date_str, "shift": "", "group_name": "Sin datos"})
        sheet_names.append(ws.title)
        sheet_num_rows[ws.title] = 0

    del wb["__template__"]
    for name in list(wb.sheetnames):
        if name not in sheet_names:
            del wb[name]

    # Restore template images (logo, etc.) to all sheets
    for sheet_name in wb.sheetnames:
        _restore_images(wb[sheet_name], template_images)

    _add_footer_images(wb, sheet_num_rows)

    output_filename = f"Reporte_Diario_{target_date.isoformat()}.xlsx"
    return wb, output_filename


def generate_daily_report(target_date: date, reports_data: dict) -> str:
    """Build daily report, save to disk, return file path."""
    ROOT = os.path.dirname(os.path.dirname(BASE_DIR))
    reports_dir = os.path.join(ROOT, "static", "daily_reports")
    os.makedirs(reports_dir, exist_ok=True)
    wb, fname = _build_workbook(target_date, reports_data)
    output_path = os.path.join(reports_dir, fname)
    wb.save(output_path)
    wb.close()
    return output_path


def generate_daily_report_bytes(target_date: date, reports_data: dict) -> io.BytesIO:
    """Build daily report into a BytesIO buffer (no disk write)."""
    wb, _ = _build_workbook(target_date, reports_data)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf


def list_daily_reports() -> List[dict]:
    ROOT = os.path.dirname(os.path.dirname(BASE_DIR))
    reports_dir = os.path.join(ROOT, "static", "daily_reports")
    reports = []
    if not os.path.isdir(reports_dir):
        return reports
    for fname in sorted(os.listdir(reports_dir), reverse=True):
        if fname.startswith("Reporte_Diario_") and fname.endswith(".xlsx"):
            path = os.path.join(reports_dir, fname)
            size = os.path.getsize(path)
            date_str = fname.replace("Reporte_Diario_", "").replace(".xlsx", "")
            reports.append({
                "filename": fname,
                "date": date_str,
                "size": size,
                "size_str": f"{size/1024:.1f} KB" if size < 1024*1024 else f"{size/(1024*1024):.1f} MB",
            })
    return reports
