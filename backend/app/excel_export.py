"""
Excel export — applies M-language–equivalent transformations:
  - Time standardization (fnEstandarizarHora)
  - Activity text cleaning / spell correction (fnLimpiarActividades)
  - Collaborator split → Trabajador N° 01–04
  - Column consolidation
  - Cross-midnight minute calculation
  - Equipment extraction
  - SKU generation
"""

import re
import math
from datetime import time, timedelta
from typing import Optional, List
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


# ─── Time standardization (fnEstandarizarHora) ──────────────────────────────

def _clean_time_text(raw: str) -> str:
    t = raw.strip().lower().strip(". ")
    replacements = {
        "cinco y media": "5:30 pm", "diez de mañana": "10:00 am",
        "medio día": "12:00 pm", "mediodía": "12:00 pm",
    }
    for k, v in replacements.items():
        t = t.replace(k, v)
    t = t.replace("ppm", "pm").replace("o", "0")
    t = t.replace("a.m.", "am").replace("p.m.", "pm")
    t = t.replace("a. m.", "am").replace("p. m.", "pm")
    t = t.replace(" am", "am").replace(" pm", "pm")
    t = t.replace(": ", ":").replace(". ", ".").replace("; ", ";")
    t = t.replace(" ", ":")  # "08 00am" → "08:00am"
    t = t.replace("y", ":").replace(".", ":").replace(";", ":")
    t = t.replace(":am", ":00am").replace(":pm", ":00pm")
    # solo número → "8:00"
    if re.fullmatch(r"\d+", t):
        t += ":00"
    # tiene am/pm sin ":"
    if ("am" in t or "pm" in t) and ":" not in t:
        t = re.sub(r"am", ":00am", t)
        t = re.sub(r"pm", ":00pm", t)
    # espacio antes de am/pm
    t = t.replace("am", " am").replace("pm", " pm").strip()
    return t


def standardize_time(val) -> Optional[str]:
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, str):
        raw = _clean_time_text(val)
        try:
            from datetime import datetime as dt
            for fmt in ("%H:%M", "%I:%M %p", "%H:%M:%S"):
                try:
                    return dt.strptime(raw, fmt).strftime("%H:%M")
                except ValueError:
                    continue
        except Exception:
            pass
    return str(val)


# ─── Activity cleaning / spell correction (fnLimpiarActividades) ──────────

SPELLING_DICT = {
    "valbula": "válvula", "balbula": "válvula",
    "valvulas": "válvulas", "balbulas": "válvulas",
    "reparasion": "reparación", "reparacion": "reparación",
    "reparasiones": "reparaciones", "reparaciones": "reparaciones",
    "manteimiento": "mantenimiento", "mantenimento": "mantenimiento",
    "limpiesa": "limpieza",
    "instalasion": "instalación", "instalacion": "instalación",
    "inspeccion": "inspección", "inspecion": "inspección",
    "electrico": "eléctrico", "mecanico": "mecánico",
    "hidraulico": "hidráulico", "bateria": "batería",
    "lubricacion": "lubricación", "lubricasion": "lubricación",
    "sitema": "sistema",
    "revision": "revisión", "revisio": "revisión",
    "calibracion": "calibración", "calibrasion": "calibración",
    "diagnostico": "diagnóstico",
    "prosedimiento": "procedimiento",
    "correccion": "corrección",
}


def clean_activity(val) -> Optional[str]:
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return None
    t = str(val)
    # remove special chars
    for ch in ("-", "✓", "*", "●", ".", "•"):
        t = t.replace(ch, " ")
    t = re.sub(r"\s+", " ", t).strip()
    # spell-check on each word
    words = t.split()
    out = []
    for w in words:
        lower = w.lower()
        if lower in SPELLING_DICT:
            # preserve original case pattern
            if w[0].isupper():
                out.append(SPELLING_DICT[lower].capitalize())
            else:
                out.append(SPELLING_DICT[lower])
        else:
            out.append(w)
    return " ".join(out)


# ─── Equipment extraction ──────────────────────────────────────────────────

def extract_equipment(equip_str: Optional[str]) -> tuple:
    """Return (equipo_interv, equipo) where equipo is text inside last (...)."""
    if not equip_str or equip_str.strip() == "":
        return ("", "")
    m = re.search(r"\(([^)]*)\)", equip_str)
    if m:
        inner = m.group(1).strip()
        return (equip_str, inner)
    return (equip_str, equip_str)


# ─── Cross-midnight minutes calculation ─────────────────────────────────────

def calc_minutes(start: Optional[str], end: Optional[str], shift: Optional[str], fallback_minutes: Optional[float]) -> Optional[float]:
    if not start or not end:
        if fallback_minutes is not None:
            try:
                return float(fallback_minutes)
            except (ValueError, TypeError):
                pass
        return None
    try:
        from datetime import datetime as dt
        s = dt.strptime(start, "%H:%M")
        e = dt.strptime(end, "%H:%M")
        diff = (e - s).total_seconds() / 60.0
        if diff < 0:
            is_night = shift and ("noche" in shift.lower() or shift.lower().startswith("n"))
            if is_night:
                diff += 720 if diff >= -720 else 1440
            else:
                diff += 720
        return round(diff, 1)
    except Exception:
        if fallback_minutes is not None:
            try:
                return float(fallback_minutes)
            except (ValueError, TypeError):
                pass
        return None


# ─── Horometer combination ──────────────────────────────────────────────────

def combine_horometer(motor, jumbo, volquetes) -> Optional[float]:
    """Return first non-None value."""
    for v in (motor, jumbo, volquetes):
        if v is not None:
            try:
                return round(float(v), 2)
            except (ValueError, TypeError):
                pass
    return None


# ─── Main Excel builder ─────────────────────────────────────────────────────

STYLES = {
    "header_fill": PatternFill("solid", fgColor="1F4E79"),
    "header_font": Font(bold=True, color="FFFFFF", size=10),
    "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
    "cell_align": Alignment(vertical="top", wrap_text=True),
    "thin_border": Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    ),
}

FINAL_COLUMNS = [
    ("Id", 8), ("SKU", 12), ("Fecha", 12),
    ("Trabajador N° 01", 22), ("Trabajador N° 02", 22),
    ("Trabajador N° 03", 22), ("Trabajador N° 04", 22),
    ("¿A que Grupo de Trabajo Perteneces?", 30),
    ("Turno de trabajo", 14),
    ("Tipo de Servicio", 28),
    ("Descripcion de Servicio", 30),
    ("Tipo de Accion", 28),
    ("Describa todas las actividades realizadas", 50),
    ("Nivel donde se trabajo", 16),
    ("Labor o lugar donde realizaste el trabajo", 30),
    ("Hora Inicio Int.", 14),
    ("Hora Termino Int.", 14),
    ("Minutos Interv.", 14),
    ("Equipo Interv.", 35),
    ("Equipo", 20),
    ("Horometro Motor", 16),
    ("Horómetro Eléctrico", 16),
    ("Horómetro de Percusión", 18),
    ("Kilometraje", 14),
    ("Tiempo Hr", 12),
]


def split_collaborators(collab_str: Optional[str]) -> List[str]:
    if not collab_str or collab_str.strip() == "":
        return []
    return [n.strip() for n in collab_str.replace(";", ",").split(",") if n.strip()]


def build_excel(reports_data: list) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "ReporteMantt"

    # ── Header row ──
    header_names = [c[0] for c in FINAL_COLUMNS]
    for col_idx, (name, width) in enumerate(FINAL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = STYLES["header_fill"]
        cell.font = STYLES["header_font"]
        cell.alignment = STYLES["header_align"]
        cell.border = STYLES["thin_border"]
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze top row ──
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(FINAL_COLUMNS))}1"

    row_num = 2
    sku_counter = 0
    for idx, rp in enumerate(reports_data, 1):
        entries = rp.get("entries", []) or [None]

        for entry in entries:
            sku_counter += 1
            if entry is None:
                entry = {}
            sku = f"Rep-{sku_counter:04d}"
            fecha = rp.get("date")

            group_name = rp.get("group_name", "")
            shift = rp.get("shift", "")
            tipo_servicio = entry.get("macroprocess", "")
            desc_servicio = entry.get("work_type", "")
            if entry.get("work_subtype"):
                desc_servicio = f"{desc_servicio} - {entry['work_subtype']}"
            tipo_accion = entry.get("action", "")
            actividades = clean_activity(entry.get("description"))
            nivel = entry.get("level", "")
            labor = entry.get("location", "")

            # Hours
            start_raw = entry.get("start_time_int")
            end_raw = entry.get("end_time_int")
            start_str = standardize_time(start_raw)
            end_str = standardize_time(end_raw)

            # Duration
            dur_raw = entry.get("duration")
            try:
                dur_min = float(dur_raw) if dur_raw is not None else None
            except (ValueError, TypeError):
                dur_min = None

            minutos_interv = calc_minutes(start_str, end_str, shift, dur_min)
            tiempo_hr = round(dur_min / 60, 2) if dur_min is not None else None

            # Equipment
            equip_raw = entry.get("equipment", "")
            equip_interv, equipo = extract_equipment(equip_raw)

            # Horometers
            horo_motor = combine_horometer(
                entry.get("horometer_motor"),
                entry.get("horometer_motor_jumbo"),
                entry.get("horometer_motor_volquetes"),
            )
            horo_electrico = entry.get("horometer_electric")
            if horo_electrico is not None:
                horo_electrico = round(float(horo_electrico), 2)
            horo_percusion = entry.get("horometer_percussion")
            if horo_percusion is not None:
                horo_percusion = round(float(horo_percusion), 2)
            km = entry.get("kilometer")
            if km is not None:
                km = round(float(km), 2)

            # Per-entry collaborators
            raw_colabs = entry.get("collaborators") or ""
            colab_parts = [c.strip() for c in raw_colabs.split(",") if c.strip()]
            trabajadores = [colab_parts[i] if i < len(colab_parts) else None for i in range(4)]

            values = [
                rp.get("id"), sku, fecha,
                trabajadores[0], trabajadores[1], trabajadores[2], trabajadores[3],
                group_name, shift,
                tipo_servicio, desc_servicio, tipo_accion,
                actividades, nivel, labor,
                start_str, end_str,
                minutos_interv,
                equip_interv, equipo,
                horo_motor, horo_electrico, horo_percusion,
                km, tiempo_hr,
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.alignment = STYLES["cell_align"]
                cell.border = STYLES["thin_border"]
                cell.font = Font(size=9)
                # Format numbers
                if col_idx in (18, 25):  # Minutos Interv, Tiempo Hr
                    cell.number_format = "#,##0.0"
                elif col_idx in (21, 22, 23, 24):  # Horometers, Km
                    cell.number_format = "#,##0.00"
                elif col_idx == 3:  # Fecha
                    cell.number_format = "DD/MM/YYYY"

            row_num += 1

    return wb
